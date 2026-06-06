from __future__ import annotations

import json
import os
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RUN_FOLDER_FORMAT = "%Y-%m-%d_%H-%M-%S"
COPYABLE_SUFFIXES = {".json", ".csv", ".html", ".txt", ".png", ".zip"}
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
BLUE_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
TIMESTAMP_SUFFIX_PATTERN = re.compile(r"_(\d{8}_\d{6}(?:_\d{1,6})?|\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})$")


def _read_bool_env(name: str, default: bool) -> bool:
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {"1", "true", "yes", "on"}


def _cleanup_report_archives(reports_root: Path, keep_latest_runs: int) -> None:
    if keep_latest_runs < 1 or not reports_root.exists():
        return

    run_directories = [path for path in reports_root.iterdir() if path.is_dir()]
    for stale_run in sorted(run_directories, key=lambda path: path.stat().st_mtime, reverse=True)[keep_latest_runs:]:
        shutil.rmtree(stale_run, ignore_errors=True)


def _cleanup_timestamped_metric_artifacts(directory: Path, canonical_files: set[str]) -> None:
    if not directory.exists():
        return

    for file_path in directory.iterdir():
        if not file_path.is_file() or file_path.name in canonical_files:
            continue
        if TIMESTAMP_SUFFIX_PATTERN.search(file_path.stem):
            file_path.unlink(missing_ok=True)


def _apply_artifact_retention_policy(root: Path) -> None:
    keep_latest_runs = max(1, int(os.getenv("REPORT_RETENTION_RUNS", "1")))
    cleanup_metric_artifacts = _read_bool_env("CLEAN_TIMESTAMPED_METRIC_ARTIFACTS", True)
    cleanup_duplicate_artifacts = _read_bool_env("CLEAN_DUPLICATE_ARTIFACTS", True)

    _cleanup_report_archives(root / "artifacts" / "reports", keep_latest_runs=keep_latest_runs)

    if not cleanup_metric_artifacts:
        cleanup_metric_artifacts = False

    if cleanup_metric_artifacts:
        _cleanup_timestamped_metric_artifacts(
            root / "artifacts" / "dspy",
            canonical_files={"dspy_results.json", "dspy_score_summary.json"},
        )
        _cleanup_timestamped_metric_artifacts(
            root / "artifacts" / "ragas",
            canonical_files={"ragas_results.json", "ragas_results.csv"},
        )

    if cleanup_duplicate_artifacts:
        duplicate_pairs = [
            (root / "artifacts" / "dspy" / "dspy_results.json", root / "reports" / "dspy_results.json"),
            (root / "artifacts" / "dspy" / "dspy_score_summary.json", root / "reports" / "dspy_score_summary.json"),
            (root / "artifacts" / "ragas" / "ragas_results.json", root / "reports" / "ragas_results.json"),
            (root / "artifacts" / "ragas" / "ragas_results.csv", root / "reports" / "ragas_results.csv"),
        ]
        for artifact_file, report_file in duplicate_pairs:
            if artifact_file.exists() and report_file.exists():
                artifact_file.unlink(missing_ok=True)


def _load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _strip_nonce(text: str | None) -> str:
    return re.sub(r"\s*\[RUN:[^\]]+\]\s*", "", str(text or "")).strip()


def _copy_tree(src: Path, dest: Path) -> None:
    if not src.exists():
        return

    for path in src.rglob("*"):
        if path.is_dir() or path.suffix.lower() not in COPYABLE_SUFFIXES:
            continue
        target = dest / path.relative_to(src)
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, target)


def _copy_file_if_exists(src: Path, dest: Path) -> None:
    if not src.exists():
        return
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)


def _generate_bridge_excel_report(bridge_report_dir: Path, ragas_csv_source: Path | None = None) -> None:
    ragas_json_path = bridge_report_dir / "ragas_results.json"
    if not ragas_json_path.exists():
        return

    ragas_payload = _load_json(ragas_json_path)
    metric_details = ragas_payload.get("metric_details", {}) if isinstance(ragas_payload, dict) else {}
    thresholds = ragas_payload.get("thresholds", {}) if isinstance(ragas_payload, dict) else {}
    executed_metrics = set(ragas_payload.get("executed_metrics", [])) if isinstance(ragas_payload, dict) else set()
    skipped_map = {
        str(item.get("metric")): str(item.get("reason"))
        for item in (ragas_payload.get("skipped_metrics", []) if isinstance(ragas_payload, dict) else [])
        if isinstance(item, dict)
    }
    summary = ragas_payload.get("summary", {}) if isinstance(ragas_payload, dict) else {}
    metric_row_counts = ragas_payload.get("metric_row_counts", {}) if isinstance(ragas_payload, dict) else {}

    ragas_csv_path = ragas_csv_source or (bridge_report_dir / "ragas_results.csv")
    ragas_csv_df = pd.DataFrame()
    if ragas_csv_path.exists():
        try:
            ragas_csv_df = pd.read_csv(ragas_csv_path)
        except Exception:
            ragas_csv_df = pd.DataFrame()
    if ragas_csv_df.empty:
        payload_rows = ragas_payload.get("rows", []) if isinstance(ragas_payload, dict) else []
        if isinstance(payload_rows, list) and payload_rows:
            ragas_csv_df = pd.DataFrame(payload_rows)

    core_metrics = ["answer_relevancy", "answer_accuracy", "response_correctness", "answer_completeness"]
    evaluator_order = [
        "answer_relevancy",
        "answer_accuracy",
        "faithfulness",
        "context_precision",
        "context_utilization",
        "context_recall",
        "context_relevance",
        "response_groundedness",
        "context_entity_recall",
        "noise_sensitivity_relevant",
        "noise_sensitivity_irrelevant",
        "response_correctness",
        "answer_completeness",
    ]

    metric_operators = {
        "answer_relevancy": ">=",
        "answer_accuracy": ">=",
        "faithfulness": ">=",
        "context_precision": ">=",
        "context_utilization": ">=",
        "context_recall": ">=",
        "context_relevance": ">=",
        "response_groundedness": ">=",
        "context_entity_recall": ">=",
        "noise_sensitivity_relevant": "<=",
        "noise_sensitivity_irrelevant": "<=",
        "response_correctness": ">=",
        "answer_completeness": ">=",
    }

    metric_score_by_question: dict[str, dict[str, Any]] = {metric_name: {} for metric_name in evaluator_order}
    for metric_name in evaluator_order:
        for item in metric_details.get(metric_name, []) or []:
            question = str(item.get("user_input") or item.get("question") or "").strip()
            if not question:
                continue
            metric_score_by_question[metric_name][question] = item.get(metric_name)

    bridge_dataset_path = bridge_report_dir / "ragas_eval_dataset.json"
    bridge_dataset_rows = []
    if bridge_dataset_path.exists():
        loaded_rows = _load_json(bridge_dataset_path)
        if isinstance(loaded_rows, list):
            bridge_dataset_rows = loaded_rows

    # Build a row-oriented view for the 4 core metrics from bridge dataset first,
    # then backfill any extra metric-only questions that were not present in dataset.
    core_rows: list[dict[str, Any]] = []
    seen_questions: set[str] = set()
    for item in bridge_dataset_rows:
        question = str(item.get("question") or item.get("user_input") or "").strip()
        if not question:
            continue
        seen_questions.add(question)

        core_row: dict[str, Any] = {
            "question": question,
            "response": item.get("response") or item.get("answer") or "",
            "ground_truth": item.get("reference") or item.get("ground_truth") or "",
        }

        metric_results: list[str] = []
        for metric_name in core_metrics:
            metric_value = metric_score_by_question.get(metric_name, {}).get(question)
            core_row[metric_name] = metric_value if metric_value not in (None, "") else "N/A"
            operator = metric_operators.get(metric_name, ">=")
            threshold = thresholds.get(metric_name)
            if threshold is None:
                metric_result = "N/A"
            else:
                metric_result = _classify_metric_result(metric_value, threshold, operator=operator)
            core_row[f"{metric_name}_result"] = metric_result
            metric_results.append(metric_result)

        if "FAIL" in metric_results:
            core_row["overall_result"] = "FAIL"
        elif "PASS" in metric_results:
            core_row["overall_result"] = "PASS"
        else:
            core_row["overall_result"] = "N/A"

        core_rows.append(core_row)

    for question in sorted({q for metric_name in core_metrics for q in metric_score_by_question.get(metric_name, {}).keys()}):
        if question in seen_questions:
            continue

        core_row = {
            "question": question,
            "response": "",
            "ground_truth": "",
        }
        metric_results = []
        for metric_name in core_metrics:
            metric_value = metric_score_by_question.get(metric_name, {}).get(question)
            core_row[metric_name] = metric_value if metric_value not in (None, "") else "N/A"
            operator = metric_operators.get(metric_name, ">=")
            threshold = thresholds.get(metric_name)
            if threshold is None:
                metric_result = "N/A"
            else:
                metric_result = _classify_metric_result(metric_value, threshold, operator=operator)
            core_row[f"{metric_name}_result"] = metric_result
            metric_results.append(metric_result)

        if "FAIL" in metric_results:
            core_row["overall_result"] = "FAIL"
        elif "PASS" in metric_results:
            core_row["overall_result"] = "PASS"
        else:
            core_row["overall_result"] = "N/A"

        core_rows.append(core_row)

    # Build a per-question score view containing all 13 evaluators + bridge evidence columns.
    question_row_map: dict[str, dict[str, Any]] = {}
    for item in bridge_dataset_rows:
        question = str(item.get("question") or item.get("user_input") or "").strip()
        if not question:
            continue
        contexts = item.get("retrieved_contexts") or item.get("contexts") or []
        citations = item.get("agent_citations") or []
        citation_quotes = item.get("agent_citation_quotes") or []
        row = {
            "id": item.get("id"),
            "question": question,
            "response": item.get("response") or item.get("answer"),
            "ground_truth": item.get("reference") or item.get("ground_truth"),
            "retrieved_chunks": "\n\n---\n\n".join(str(c) for c in contexts if c),
            "citations": "\n".join(str(c) for c in citations if c),
            "citation_quotes": "\n".join(str(c) for c in citation_quotes if c),
        }
        for metric_name in evaluator_order:
            row[metric_name] = None
        question_row_map[question] = row

    for metric_name in evaluator_order:
        for item in metric_details.get(metric_name, []) or []:
            question = str(item.get("user_input") or item.get("question") or "").strip()
            if not question:
                continue
            row = question_row_map.setdefault(
                question,
                {
                    "id": item.get("id"),
                    "question": question,
                    "response": item.get("response"),
                    "ground_truth": item.get("reference") or item.get("ground_truth"),
                    "retrieved_chunks": "\n\n---\n\n".join(str(c) for c in (item.get("retrieved_contexts") or [])),
                    "citations": "",
                    "citation_quotes": "",
                    **{name: None for name in evaluator_order},
                },
            )
            if not row.get("response") and item.get("response"):
                row["response"] = item.get("response")
            if not row.get("ground_truth") and (item.get("reference") or item.get("ground_truth")):
                row["ground_truth"] = item.get("reference") or item.get("ground_truth")
            row[metric_name] = item.get(metric_name)

    all_evaluator_rows = list(question_row_map.values())

    # __PATCH_EXCEL_ROWS_ALIAS__
    # Auto-injected alias so the KPI / summary / metric_explanations sections work.
    # `rows` was referenced 11 times in this function but never assigned.
    # `all_evaluator_rows` (built above) has the exact shape they expect:
    #   id, question, response, ground_truth, retrieved_chunks, citations,
    #   citation_quotes, plus per-metric score columns.
    rows = all_evaluator_rows

    # Enrich `rows` with latency / token / citation info from the bridge dataset
    # (keyed by question) so KPI averages aren't all zero.
    _bridge_by_q = {
        str(_it.get("question") or _it.get("user_input") or "").strip(): _it
        for _it in bridge_dataset_rows
        if isinstance(_it, dict)
    }
    for _r in rows:
        _q = str(_r.get("question") or "").strip()
        _src = _bridge_by_q.get(_q, {}) if _q else {}
        if "latency_seconds" not in _r:
            _r["latency_seconds"] = _src.get("latency_seconds")
        if "total_tokens" not in _r:
            _tu = _src.get("token_usage") or {}
            _r["total_tokens"] = _tu.get("total_tokens") if isinstance(_tu, dict) else None
        # Make sure 'citations' is a non-empty-aware truthy/falsey value
        if _r.get("citations") is None:
            _r["citations"] = ""

    # `source_test_suite` is referenced in the KPI section. Derive from the
    # first bridge dataset row if available; otherwise from ragas_payload.
    try:
        source_test_suite
    except NameError:
        _src_suite = ""
        if bridge_dataset_rows and isinstance(bridge_dataset_rows[0], dict):
            _src_suite = str(bridge_dataset_rows[0].get("source_test_suite") or "")
        if not _src_suite and isinstance(ragas_payload, dict):
            _src_suite = str(ragas_payload.get("source_test_suite") or "")
        source_test_suite = _src_suite or "bridge_run"


    threshold_rows = []
    for metric_name in evaluator_order:
        operator = "<=" if metric_name in {"noise_sensitivity_relevant", "noise_sensitivity_irrelevant"} else ">="
        threshold_rows.append(
            {
                "evaluator": metric_name,
                "operator": operator,
                "threshold": thresholds.get(metric_name) if metric_name in executed_metrics else None,
                "executed": "YES" if metric_name in executed_metrics else "NO",
                "average_score": summary.get(metric_name),
                "rows_evaluated": metric_row_counts.get(metric_name, 0),
                "skip_reason": skipped_map.get(metric_name, ""),
            }
        )

    excel_path = bridge_report_dir / "Bridge_Evaluation_Report.xlsx"

    # =============================
    # ✅ TEST SUMMARY — KPI SECTION
    # =============================
    total_cases = len(rows)
    answered = sum(1 for r in rows if r.get("response"))
    refusals = sum(1 for r in rows if not r.get("response"))
    with_contexts = sum(1 for r in rows if r.get("retrieved_chunks"))
    with_gt = sum(1 for r in rows if r.get("ground_truth"))
    with_citations = sum(1 for r in rows if r.get("citations"))

    avg_latency = round(
        sum((r.get("latency_seconds") or 0) for r in rows) / max(total_cases, 1), 2
    )
    avg_tokens = int(
        sum((r.get("total_tokens") or 0) for r in rows) / max(total_cases, 1)
    )

    kpi_rows = [
        {"metric": "Test Suite", "value": source_test_suite or "N/A"},
        {"metric": "Input File", "value": (source_test_suite or "N/A") + ".json"},
        {"metric": "Total Test Cases", "value": total_cases},
        {"metric": "Answered (real response)", "value": answered},
        {"metric": "Refusals / Errors", "value": refusals},
        {"metric": "With Contexts", "value": with_contexts},
        {"metric": "With Ground Truth", "value": with_gt},
        {"metric": "With Citations", "value": with_citations},
        {"metric": "Avg Latency (s)", "value": avg_latency},
        {"metric": "Avg Total Tokens", "value": avg_tokens},
    ]

    # =============================
    # ✅ TEST SUMMARY — EVALUATOR COUNTS
    # =============================
    summary_rows = []
    for metric in evaluator_order:
        total = total_cases
        threshold = thresholds.get(metric)
        operator = metric_operators.get(metric, ">=")

        passed = failed = skipped = 0

        for r in rows:
            val = r.get(metric)
            if val in (None, ""):
                skipped += 1
                continue
            try:
                val = float(val)
                th = float(threshold)
                if operator == "<=":
                    if val <= th:
                        passed += 1
                    else:
                        failed += 1
                else:
                    if val >= th:
                        passed += 1
                    else:
                        failed += 1
            except:
                skipped += 1

        pass_rate = f"{(passed / max(total, 1)) * 100:.0f}%" if total else "N/A"

        if passed == 0 and failed == 0:
            pass_rate = "N/A"

        summary_rows.append({
            "evaluator": metric,
            "total": total,
            "pass": passed,
            "fail": failed,
            "skipped": skipped,
            "pass_rate": pass_rate,
        })

    # =============================
    # ✅ METRIC EXPLANATIONS — PER ROW × PER METRIC
    # =============================
    metric_descriptions = {
        "answer_relevancy": "Measures if the response directly addresses the question.",
        "answer_accuracy": "Compares factual claims in response vs ground truth.",
        "faithfulness": "Checks if response claims are supported by retrieved contexts.",
        "context_precision": "Measures if retrieved contexts are relevant and precise.",
        "context_utilization": "Checks if the response uses retrieved contexts effectively.",
        "context_recall": "Measures if contexts cover all claims in ground truth.",
        "context_relevance": "Evaluates if retrieved contexts are relevant to the question.",
        "response_groundedness": "Checks if every claim in response is grounded in contexts.",
        "context_entity_recall": "Compares named entities between contexts and ground truth.",
        "noise_sensitivity_relevant": "Measures if adding relevant noise changes the answer.",
        "noise_sensitivity_irrelevant": "Measures if adding irrelevant noise changes the answer.",
        "response_correctness": "Overall correctness combining factual accuracy and semantic similarity.",
        "answer_completeness": "Rates how completely the response covers all aspects (1-5 scale).",
    }

    metric_explanations = []

    for r in rows:
        row_id = r.get("id", "")
        question = r.get("question", "")

        for metric in evaluator_order:
            score = r.get(metric)
            threshold = thresholds.get(metric)
            operator = metric_operators.get(metric, ">=")
            description = metric_descriptions.get(metric, "")

            # Determine result + explanation
            if score in (None, ""):
                result = "SKIPPED"
                explanation = "Metric was skipped or returned no score."
            else:
                try:
                    score_f = float(score)
                    th_f = float(threshold)

                    if operator == "<=":
                        if score_f <= th_f:
                            result = "PASS"
                            explanation = f"Score {score_f:.4f} <= threshold {th_f}. Noise impact is acceptable."
                        else:
                            result = "FAIL"
                            explanation = f"Score {score_f:.4f} > threshold {th_f}. Response is too sensitive to noise."
                    else:
                        if score_f >= th_f:
                            result = "PASS"

                            # Custom messages
                            if metric == "faithfulness":
                                explanation = f"Score {score_f:.4f} >= threshold {th_f}. Response is grounded in source contexts."
                            elif metric == "context_utilization":
                                explanation = f"Score {score_f:.4f} >= threshold {th_f}. Response effectively uses the retrieved contexts."
                            elif metric == "response_groundedness":
                                explanation = f"Score {score_f:.4f} >= threshold {th_f}. All response claims are grounded in contexts."
                            else:
                                explanation = f"Score {score_f:.4f} >= threshold {th_f}. Metric passed."
                        else:
                            result = "FAIL"

                            if metric == "answer_accuracy":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Response has different facts or wording than ground truth."
                            elif metric == "context_precision":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Retrieved contexts contain noisy or irrelevant information."
                            elif metric == "context_recall":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Contexts are missing key information from ground truth."
                            elif metric == "context_relevance":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Retrieved contexts are not relevant to the question."
                            elif metric == "response_correctness":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Response is incorrect or semantically different from ground truth."
                            elif metric == "faithfulness":
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Response contains claims not found in contexts (possible hallucination)."
                            else:
                                explanation = f"Score {score_f:.4f} < threshold {th_f}. Metric below acceptable level."
                except:
                    result = "SKIPPED"
                    explanation = "Could not evaluate score."

            metric_explanations.append({
                "id": row_id,
                "question": question,
                "metric": metric,
                "score": score if score not in (None, "") else "",
                "threshold": threshold,
                "result": result,
                "explanation": explanation,
                "metric_description": description,
            })

    # ✅ WRITE EXCEL
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Test_Summary", index=False)
        pd.DataFrame(metric_explanations).to_excel(writer, sheet_name="Metric_Explanations", index=False)

        pd.DataFrame(
            core_rows,
            columns=[
                "question", "response", "ground_truth",
                "answer_relevancy", "answer_relevancy_result",
                "answer_accuracy", "answer_accuracy_result",
                "response_correctness", "response_correctness_result",
                "answer_completeness", "answer_completeness_result",
                "overall_result"
            ]
        ).to_excel(writer, sheet_name="Core_4_Metrics", index=False)

        pd.DataFrame(
            all_evaluator_rows,
            columns=[
                "id", "question", "response", "ground_truth",
                "retrieved_chunks", "citations", "citation_quotes",
                *evaluator_order
            ]
        ).to_excel(writer, sheet_name="All_13_Evaluators", index=False)

        pd.DataFrame(
            threshold_rows,
            columns=[
                "evaluator", "operator", "threshold", "executed",
                "average_score", "rows_evaluated", "skip_reason"
            ]
        ).to_excel(writer, sheet_name="Thresholds", index=False)

        ragas_csv_df.to_excel(writer, sheet_name="RAGAS_Results_CSV", index=False)

    workbook = load_workbook(excel_path)
    for sheet in workbook.worksheets:
        if sheet.max_row >= 1:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_cells in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 14), 80)

    if "Thresholds" in workbook.sheetnames:
        threshold_sheet = workbook["Thresholds"]
        headers = {
            str(threshold_sheet.cell(row=1, column=index).value or ""): index
            for index in range(1, threshold_sheet.max_column + 1)
        }
        for row_index in range(2, threshold_sheet.max_row + 1):
            executed_cell = threshold_sheet.cell(row=row_index, column=headers.get("executed", 0))
            if executed_cell.value == "YES":
                executed_cell.fill = GREEN_FILL
            else:
                executed_cell.fill = YELLOW_FILL

    if "All_13_Evaluators" in workbook.sheetnames:
        eval_sheet = workbook["All_13_Evaluators"]
        headers = {
            str(eval_sheet.cell(row=1, column=index).value or ""): index
            for index in range(1, eval_sheet.max_column + 1)
        }
        wrap_headers = ["question", "response", "ground_truth", "retrieved_chunks", "citations", "citation_quotes"]
        for row_index in range(2, eval_sheet.max_row + 1):
            for wrap_header in wrap_headers:
                if wrap_header in headers:
                    eval_sheet.cell(row=row_index, column=headers[wrap_header]).alignment = Alignment(
                        wrap_text=True,
                        vertical="top",
                    )

            for metric_name in evaluator_order:
                if metric_name not in headers:
                    continue
                score_cell = eval_sheet.cell(row=row_index, column=headers[metric_name])
                score_value = score_cell.value
                if score_value in (None, ""):
                    score_cell.fill = YELLOW_FILL
                    continue

                threshold_value = thresholds.get(metric_name)
                if threshold_value is None:
                    continue

                try:
                    numeric_score = float(score_value)
                    numeric_threshold = float(threshold_value)
                    if metric_name in {"noise_sensitivity_relevant", "noise_sensitivity_irrelevant"}:
                        score_cell.fill = GREEN_FILL if numeric_score <= numeric_threshold else RED_FILL
                    else:
                        score_cell.fill = GREEN_FILL if numeric_score >= numeric_threshold else RED_FILL
                except Exception:
                    score_cell.fill = YELLOW_FILL

    if "Core_4_Metrics" in workbook.sheetnames:
        core_sheet = workbook["Core_4_Metrics"]
        headers = {str(core_sheet.cell(row=1, column=index).value or ""): index for index in range(1, core_sheet.max_column + 1)}
        for row_index in range(2, core_sheet.max_row + 1):
            for wrap_header in ["question", "response", "ground_truth"]:
                if wrap_header in headers:
                    core_sheet.cell(row=row_index, column=headers[wrap_header]).alignment = Alignment(wrap_text=True, vertical="top")

            for metric_name in core_metrics:
                if metric_name in headers:
                    score_cell = core_sheet.cell(row=row_index, column=headers[metric_name])
                    score_value = score_cell.value
                    if score_value in (None, ""):
                        score_cell.fill = YELLOW_FILL
                    else:
                        threshold_value = thresholds.get(metric_name)
                        if threshold_value is None:
                            score_cell.fill = YELLOW_FILL
                        else:
                            try:
                                numeric_score = float(score_value)
                                numeric_threshold = float(threshold_value)
                                if metric_operators.get(metric_name) == "<=":
                                    score_cell.fill = GREEN_FILL if numeric_score <= numeric_threshold else RED_FILL
                                else:
                                    score_cell.fill = GREEN_FILL if numeric_score >= numeric_threshold else RED_FILL
                            except Exception:
                                score_cell.fill = YELLOW_FILL

                result_header = f"{metric_name}_result"
                if result_header in headers:
                    result_cell = core_sheet.cell(row=row_index, column=headers[result_header])
                    if result_cell.value == "PASS":
                        result_cell.fill = GREEN_FILL
                    elif result_cell.value == "FAIL":
                        result_cell.fill = RED_FILL
                    else:
                        result_cell.fill = YELLOW_FILL

            if "overall_result" in headers:
                overall_cell = core_sheet.cell(row=row_index, column=headers["overall_result"])
                if overall_cell.value == "PASS":
                    overall_cell.fill = GREEN_FILL
                elif overall_cell.value == "FAIL":
                    overall_cell.fill = RED_FILL
                else:
                    overall_cell.fill = YELLOW_FILL

    workbook.save(excel_path)


def _prune_ui_e2e_report_columns(ui_report_path: Path) -> None:
    """Remove all RAGAS metric columns from UI E2E report — RAGAS is only for Bridge mode."""
    if not ui_report_path.exists():
        return

    workbook = load_workbook(ui_report_path)
    if "Test Results" not in workbook.sheetnames:
        workbook.save(ui_report_path)
        return

    sheet = workbook["Test Results"]
    headers = [str(sheet.cell(row=1, column=column).value or "") for column in range(1, sheet.max_column + 1)]
    header_to_index = {header: index + 1 for index, header in enumerate(headers)}

    # All RAGAS columns — remove them entirely from UI E2E mode
    ragas_metric_names = [
        "ragas_answer_relevancy",
        "ragas_answer_accuracy",
        "ragas_faithfulness",
        "ragas_context_precision",
        "ragas_context_utilization",
        "ragas_context_recall",
        "ragas_context_relevance",
        "ragas_response_groundedness",
        "ragas_context_entity_recall",
        "ragas_noise_sensitivity_relevant",
        "ragas_noise_sensitivity_irrelevant",
        "ragas_response_correctness",
        "ragas_answer_completeness",
    ]

    delete_columns: list[int] = []
    for metric_name in ragas_metric_names:
        result_header = f"{metric_name}_result"
        threshold_header_candidates = [f"{metric_name}_pass_if>=", f"{metric_name}_pass_if<="]
        # Remove metric, result, and threshold columns
        for header_name in [metric_name, *threshold_header_candidates, result_header]:
            column_index = header_to_index.get(header_name)
            if column_index is not None:
                delete_columns.append(column_index)

    for column_index in sorted(set(delete_columns), reverse=True):
        sheet.delete_cols(column_index, 1)

    workbook.save(ui_report_path)


def _sync_canonical_outputs_to_reports(root: Path, report_mode: str = "all") -> None:
    """Populate report folders based on *report_mode*.

    report_mode values:
      "bridge"  – only generate reports/bridge/ outputs (RAGAS / bridge pipeline)
      "ui_e2e"  – only generate reports/ui_e2e/ outputs (UI + DSPy pipeline)
      "all"     – generate both (default, backward-compatible)
    """
    reports_dir = root / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    ui_report_dir = reports_dir / "ui_e2e"
    bridge_report_dir = reports_dir / "bridge"
    temp_latest_report = reports_dir / "Latest_Report.xlsx"

    generate_ui_e2e = report_mode in {"ui_e2e", "all"}
    generate_bridge = report_mode in {"bridge", "all"}

    if generate_ui_e2e:
        ui_report_dir.mkdir(parents=True, exist_ok=True)
        # Remove stale sample workbooks so ui_e2e contains only canonical outputs.
        for stale_sample in ui_report_dir.glob("*_SAMPLE.xlsx"):
            stale_sample.unlink(missing_ok=True)
        # All UI/E2E outputs go only into reports/ui_e2e/ — nothing at the reports/ root.
        ui_report_path = ui_report_dir / "Latest_Report_UI_E2E.xlsx"
        _copy_file_if_exists(temp_latest_report, ui_report_path)
        _prune_ui_e2e_report_columns(ui_report_path)
        # pytest HTML — overall test pass/fail summary for this run.
        _copy_file_if_exists(root / "artifacts" / "reports" / "pytest_report.html", ui_report_dir / "pytest_report.html")
        # Playwright trace ZIPs — per-test screenshots, clicks, DOM snapshots (open with: playwright show-trace <file>).
        traces_src = root / "artifacts" / "ui_runs"
        traces_dst = ui_report_dir / "playwright_traces"
        if traces_src.exists():
            traces_dst.mkdir(parents=True, exist_ok=True)
            for trace_zip in traces_src.rglob("trace.zip"):
                dest = traces_dst / trace_zip.parent.name / "trace.zip"
                dest.parent.mkdir(parents=True, exist_ok=True)
                _copy_file_if_exists(trace_zip, dest)
        _copy_file_if_exists(root / "artifacts" / "dspy" / "dspy_results.json", ui_report_dir / "dspy_results.json")
        _copy_file_if_exists(root / "artifacts" / "dspy" / "dspy_score_summary.json", ui_report_dir / "dspy_score_summary.json")

    if generate_bridge:
        bridge_report_dir.mkdir(parents=True, exist_ok=True)
        # All bridge outputs go only into reports/bridge/ — nothing at the reports/ root.
        # pytest HTML — overall pass/fail for the ragas/bridge test run.
        _copy_file_if_exists(root / "artifacts" / "reports" / "pytest_report.html", bridge_report_dir / "pytest_report.html")
        _copy_file_if_exists(root / "artifacts" / "ragas" / "ragas_results.json", bridge_report_dir / "ragas_results.json")
        _copy_file_if_exists(root / "data" / "ragas_eval_dataset.json", bridge_report_dir / "ragas_eval_dataset.json")
        _generate_bridge_excel_report(
            bridge_report_dir,
            ragas_csv_source=root / "artifacts" / "ragas" / "ragas_results.csv",
        )
        # Raw CSV is now embedded into the Excel workbook as a dedicated sheet.
        (bridge_report_dir / "ragas_results.csv").unlink(missing_ok=True)

    # Always remove intermediate root workbook, even for bridge-only runs.
    temp_latest_report.unlink(missing_ok=True)


def _load_ui_artifacts(ui_dir: Path) -> dict[str, dict[str, Any]]:
    artifacts: dict[str, dict[str, Any]] = {}
    if not ui_dir.exists():
        return artifacts

    for artifact_file in sorted(ui_dir.glob("*.json")):
        payload = _load_json(artifact_file)
        if payload:
            test_id = str(payload.get("id") or payload.get("test_id") or artifact_file.stem)
            artifacts[test_id] = payload
    return artifacts


def _load_test_cases(project_root: Path) -> dict[str, dict[str, Any]]:
    cases_path = project_root / "data" / "test_cases.json"
    if not cases_path.exists():
        return {}
    try:
        cases = json.loads(cases_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {str(item.get("id")): item for item in cases if item.get("id")}


def _read_float(raw_value: Any, default: float) -> float:
    try:
        return float(raw_value)
    except Exception:
        return float(default)


def _build_metric_thresholds(ragas_results: dict[str, Any]) -> dict[str, float]:
    ragas_thresholds = ragas_results.get("thresholds", {}) if isinstance(ragas_results, dict) else {}
    return {
        "dspy_total_score": _read_float(os.getenv("DSPY_MIN_SCORE", "0.40"), 0.40),
        "dspy_keyword_score": _read_float(os.getenv("DSPY_KEYWORD_THRESHOLD", "1.0"), 1.0),
        "dspy_fallback_score": _read_float(os.getenv("DSPY_FALLBACK_THRESHOLD", "1.0"), 1.0),
        "dspy_format_score": _read_float(os.getenv("DSPY_FORMAT_THRESHOLD", "1.0"), 1.0),
        "dspy_pdf_grounding_score": _read_float(os.getenv("DSPY_PDF_GROUNDING_THRESHOLD", "1.0"), 1.0),
        "ragas_answer_relevancy": _read_float(
            ragas_thresholds.get("answer_relevancy", os.getenv("RAGAS_ANSWER_RELEVANCY_THRESHOLD", "0.65")),
            0.65,
        ),
        "ragas_answer_accuracy": _read_float(
            ragas_thresholds.get("answer_accuracy", os.getenv("RAGAS_ANSWER_ACCURACY_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_faithfulness": _read_float(
            ragas_thresholds.get("faithfulness", os.getenv("RAGAS_FAITHFULNESS_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_context_precision": _read_float(
            ragas_thresholds.get("context_precision", os.getenv("RAGAS_CONTEXT_PRECISION_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_context_utilization": _read_float(
            ragas_thresholds.get("context_utilization", os.getenv("RAGAS_CONTEXT_UTILIZATION_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_context_recall": _read_float(
            ragas_thresholds.get("context_recall", os.getenv("RAGAS_CONTEXT_RECALL_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_context_relevance": _read_float(
            ragas_thresholds.get("context_relevance", os.getenv("RAGAS_CONTEXT_RELEVANCE_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_response_groundedness": _read_float(
            ragas_thresholds.get("response_groundedness", os.getenv("RAGAS_RESPONSE_GROUNDEDNESS_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_context_entity_recall": _read_float(
            ragas_thresholds.get("context_entity_recall", os.getenv("RAGAS_CONTEXT_ENTITY_RECALL_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_noise_sensitivity_relevant": _read_float(
            ragas_thresholds.get("noise_sensitivity_relevant", os.getenv("RAGAS_NOISE_SENSITIVITY_RELEVANT_THRESHOLD", "0.30")),
            0.30,
        ),
        "ragas_noise_sensitivity_irrelevant": _read_float(
            ragas_thresholds.get("noise_sensitivity_irrelevant", os.getenv("RAGAS_NOISE_SENSITIVITY_IRRELEVANT_THRESHOLD", "0.30")),
            0.30,
        ),
        "ragas_response_correctness": _read_float(
            ragas_thresholds.get("response_correctness", os.getenv("RAGAS_RESPONSE_CORRECTNESS_THRESHOLD", "0.70")),
            0.70,
        ),
        "ragas_answer_completeness": _read_float(
            ragas_thresholds.get("answer_completeness", os.getenv("RAGAS_ANSWER_COMPLETENESS_THRESHOLD", "3.0")),
            3.0,
        ),
    }


def _build_metric_operators() -> dict[str, str]:
    return {
        "dspy_total_score": ">=",
        "dspy_keyword_score": ">=",
        "dspy_fallback_score": ">=",
        "dspy_format_score": ">=",
        "dspy_pdf_grounding_score": ">=",
        "ragas_answer_relevancy": ">=",
        "ragas_answer_accuracy": ">=",
        "ragas_faithfulness": ">=",
        "ragas_context_precision": ">=",
        "ragas_context_utilization": ">=",
        "ragas_context_recall": ">=",
        "ragas_context_relevance": ">=",
        "ragas_response_groundedness": ">=",
        "ragas_context_entity_recall": ">=",
        "ragas_noise_sensitivity_relevant": "<=",
        "ragas_noise_sensitivity_irrelevant": "<=",
        "ragas_response_correctness": ">=",
        "ragas_answer_completeness": ">=",
    }


def _classify_metric_result(value: Any, threshold: float, operator: str = ">=") -> str:
    if value in (None, ""):
        return "N/A"
    try:
        numeric_value = float(value)
        numeric_threshold = float(threshold)
        if operator == "<=":
            return "PASS" if numeric_value <= numeric_threshold else "FAIL"
        return "PASS" if numeric_value >= numeric_threshold else "FAIL"
    except Exception:
        return "N/A"


def _build_results_rows(
    project_root: Path,
    ui_artifacts: dict[str, dict[str, Any]],
    dspy_results: dict[str, Any],
    ragas_results: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, float]]:
    dspy_rows = dspy_results.get("results", []) if isinstance(dspy_results, dict) else []
    dspy_by_id = {str(row.get("id")): row for row in dspy_rows if row.get("id") is not None}
    ragas_rows = ragas_results.get("rows", []) if isinstance(ragas_results, dict) else []
    ragas_by_id = {str(row.get("id")): row for row in ragas_rows if row.get("id") is not None}
    ragas_by_question = {
        _strip_nonce(row.get("question") or row.get("user_input")): row
        for row in ragas_rows
        if _strip_nonce(row.get("question") or row.get("user_input"))
    }
    test_cases_by_id = _load_test_cases(project_root)
    metric_thresholds = _build_metric_thresholds(ragas_results)
    metric_operators = _build_metric_operators()
    metric_order = [
        "dspy_total_score",
        "dspy_keyword_score",
        "dspy_fallback_score",
        "dspy_format_score",
        "dspy_pdf_grounding_score",
        "ragas_answer_relevancy",
        "ragas_answer_accuracy",
        "ragas_faithfulness",
        "ragas_context_precision",
        "ragas_context_utilization",
        "ragas_context_recall",
        "ragas_context_relevance",
        "ragas_response_groundedness",
        "ragas_context_entity_recall",
        "ragas_noise_sensitivity_relevant",
        "ragas_noise_sensitivity_irrelevant",
        "ragas_response_correctness",
        "ragas_answer_completeness",
    ]
    metric_counters = {metric: {"PASS": 0, "FAIL": 0, "N/A": 0} for metric in metric_order}

    results_rows: list[dict[str, Any]] = []
    failures_only_rows: list[dict[str, Any]] = []

    for test_id, payload in ui_artifacts.items():
        answer_text = str(payload.get("answer_text") or payload.get("answer") or "").strip()
        base_prompt = _strip_nonce(payload.get("base_prompt") or payload.get("prompt"))
        run_id = str(payload.get("run_id") or "").strip()
        nonce_token = str(payload.get("nonce_token") or "").strip()
        prompt_sent = str(payload.get("prompt_sent") or "")
        run_id_visible_in_prompt = bool(payload.get("run_id_visible_in_prompt", False))
        run_id_nonce_present = bool(nonce_token and nonce_token in prompt_sent)
        dspy_row = dspy_by_id.get(test_id, {})
        ragas_row = ragas_by_id.get(test_id) or ragas_by_question.get(base_prompt, {})
        test_case = test_cases_by_id.get(test_id, {})
        deterministic_scores = dspy_row.get("deterministic_scores") or {}

        expected_pdfs = dspy_row.get("expected_pdfs") or payload.get("expected_pdfs") or test_case.get("expected_pdfs") or []
        matched_pdfs = dspy_row.get("matched_pdfs") or payload.get("matched_pdfs") or []
        ground_truths = dspy_row.get("ground_truths") or []
        ground_truth = payload.get("ground_truth") or (ground_truths[0] if ground_truths else "") or test_case.get("ground_truth") or ""

        metric_values = {
            "dspy_total_score": deterministic_scores.get("total", ""),
            "dspy_keyword_score": deterministic_scores.get("keyword_presence", ""),
            "dspy_fallback_score": deterministic_scores.get("fallback_detection", ""),
            "dspy_format_score": deterministic_scores.get("formatting_constraints", ""),
            "dspy_pdf_grounding_score": deterministic_scores.get("pdf_grounding", ""),
            "ragas_answer_relevancy": ragas_row.get("answer_relevancy", ""),
            "ragas_answer_accuracy": ragas_row.get("answer_accuracy", ""),
            "ragas_faithfulness": ragas_row.get("faithfulness", ""),
            "ragas_context_precision": ragas_row.get("context_precision", ""),
            "ragas_context_utilization": ragas_row.get("context_utilization", ""),
            "ragas_context_recall": ragas_row.get("context_recall", ""),
            "ragas_context_relevance": ragas_row.get("context_relevance", ""),
            "ragas_response_groundedness": ragas_row.get("response_groundedness", ""),
            "ragas_context_entity_recall": ragas_row.get("context_entity_recall", ""),
            "ragas_noise_sensitivity_relevant": ragas_row.get("noise_sensitivity_relevant", ""),
            "ragas_noise_sensitivity_irrelevant": ragas_row.get("noise_sensitivity_irrelevant", ""),
            "ragas_response_correctness": ragas_row.get("response_correctness", ""),
            "ragas_answer_completeness": ragas_row.get("answer_completeness", ""),
        }

        issues = list(dspy_row.get("issues", []) or [])
        if not answer_text:
            issues.append("Bot answer was empty.")
        if dspy_row.get("source_match_status") == "matched_unexpected_pdf":
            issues.append("Observed evidence matched an unexpected PDF.")

        row: dict[str, Any] = {
            "test_id": test_id,
            "prompt": base_prompt or str(test_case.get("prompt") or payload.get("prompt") or ""),
            "run_id": run_id,
            "run_id_visible_in_prompt": "YES" if run_id_visible_in_prompt else "NO",
            "run_id_nonce_present": "YES" if run_id_nonce_present else "NO",
            "answer_text": answer_text,
            "ground_truth": str(ground_truth),
        }

        metric_failures: list[str] = []
        for metric_name in metric_order:
            threshold = metric_thresholds[metric_name]
            operator = metric_operators.get(metric_name, ">=")
            metric_value = metric_values.get(metric_name, "")
            metric_result = _classify_metric_result(metric_value, threshold, operator=operator)
            metric_counters[metric_name][metric_result] += 1
            row[metric_name] = metric_value
            row[f"{metric_name}_pass_if{operator}"] = threshold
            row[f"{metric_name}_result"] = metric_result
            if metric_result == "FAIL":
                comparison_hint = ">" if operator == "<=" else "<"
                metric_failures.append(f"{metric_name} outside threshold ({metric_value} {comparison_hint} {threshold})")

        deduped_issues = list(dict.fromkeys(issue for issue in [*issues, *metric_failures] if issue))
        status = "FAIL" if deduped_issues else "PASS"
        failure_reason = " | ".join(deduped_issues)

        row["final_status"] = status
        row["status"] = status
        row["expected_pdfs"] = ", ".join(expected_pdfs) if expected_pdfs else ""
        row["matched_pdfs"] = ", ".join(matched_pdfs) if matched_pdfs else ""
        row["artifact_json_path"] = str(project_root / "artifacts" / "ui_runs" / f"{test_id}.json")
        row["failure_screenshot_path"] = str(project_root / "artifacts" / "ui_runs" / test_id / "screenshot.png")
        row["failure_trace_path"] = str(project_root / "artifacts" / "ui_runs" / test_id / "trace.zip")
        results_rows.append(row)

        if status == "FAIL":
            failures_only_rows.append(
                {
                    "test_id": test_id,
                    "failure_reason": failure_reason,
                    "answer_text_excerpt": answer_text[:250],
                    "artifact_json_path": row["artifact_json_path"],
                    "failure_screenshot_path": row["failure_screenshot_path"],
                    "failure_trace_path": row["failure_trace_path"],
                }
            )

    test_summary_rows = [
        {
            "evaluator": metric_name,
            "pass_rule": f"score {metric_operators.get(metric_name, '>=')} threshold",
            "pass_threshold": metric_thresholds[metric_name],
            "PASS": metric_counters[metric_name]["PASS"],
            "FAIL": metric_counters[metric_name]["FAIL"],
            "N/A": metric_counters[metric_name]["N/A"],
        }
        for metric_name in metric_order
    ]

    return results_rows, failures_only_rows, test_summary_rows, metric_thresholds


def _build_pdf_coverage_rows(project_root: Path, results_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    registry_path = project_root / "data" / "pdf_registry.json"
    registry = []
    if registry_path.exists():
        registry = json.loads(registry_path.read_text(encoding="utf-8"))

    coverage_rows: list[dict[str, Any]] = []
    for record in registry:
        pdf_id = str(record.get("pdf_id") or "")
        pdf_name = str(record.get("pdf_name") or pdf_id)
        matching_rows = [row for row in results_rows if pdf_id and pdf_id in row.get("expected_pdfs", "")]
        total_tests = len(matching_rows)
        coverage_rows.append(
            {
                "pdf_name": pdf_name,
                "total_tests": total_tests,
                "tested": "YES" if total_tests > 0 else "NO",
            }
        )

    return coverage_rows


def _build_summary_rows(results_rows: list[dict[str, Any]], pdf_coverage_rows: list[dict[str, Any]], execution_timestamp: str) -> list[dict[str, Any]]:
    passed = sum(1 for row in results_rows if row.get("final_status", row.get("status")) == "PASS")
    failed = sum(1 for row in results_rows if row.get("final_status", row.get("status")) == "FAIL")
    total = len(results_rows)
    pass_rate = round((passed / total) * 100, 2) if total else 0.0
    run_id_captured = sum(1 for row in results_rows if str(row.get("run_id") or "").strip())
    run_id_visible = sum(1 for row in results_rows if row.get("run_id_visible_in_prompt") == "YES")
    run_id_nonce_verified = sum(1 for row in results_rows if row.get("run_id_nonce_present") == "YES")
    return [
        {"metric": "total_tests", "value": total},
        {"metric": "passed", "value": passed},
        {"metric": "failed", "value": failed},
        {"metric": "pass_rate_percent", "value": pass_rate},
        {"metric": "run_id_captured_cases", "value": run_id_captured},
        {"metric": "run_id_visible_prompt_cases", "value": run_id_visible},
        {"metric": "run_id_nonce_verified_cases", "value": run_id_nonce_verified},
        {"metric": "untested_pdfs", "value": sum(1 for row in pdf_coverage_rows if row["tested"] == "NO")},
        {"metric": "execution_timestamp", "value": execution_timestamp},
        {"metric": "traceability", "value": "Each row maps to artifacts/ui_runs/<test_id>.json for the real chatbot answer."},
    ]


def _build_legend_rows(metric_thresholds: dict[str, float]) -> list[dict[str, Any]]:
    return [
        {"field": "prompt", "description": "The user question sent to the chatbot.", "importance": "Helps reviewers see exactly what was tested."},
        {"field": "run_id", "description": "Unique run identifier captured for this test case.", "importance": "Confirms request-level traceability for this execution."},
        {"field": "run_id_visible_in_prompt", "description": "Whether run ID was intentionally appended to the sent prompt.", "importance": "Verifies the run-id visibility mode used during UI execution."},
        {"field": "run_id_nonce_present", "description": "Whether nonce token was found inside prompt_sent.", "importance": "Validates that run-id nonce was actually injected when visibility is enabled."},
        {"field": "answer_text", "description": "The real chatbot response captured from the UI.", "importance": "This is the business-visible answer for QA and audit review."},
        {"field": "ground_truth", "description": "The expected reference answer or expected behavior for the test case.", "importance": "Lets reviewers compare actual output with the intended expected outcome."},
        {"field": "final_status", "description": "Overall final test decision.", "importance": "Lets managers quickly see whether a test case passed or failed."},
        {"field": "dspy_total_score", "description": f"Combined deterministic DSPy score. Current threshold: {metric_thresholds['dspy_total_score']}", "importance": "Primary fast QA gate for structure, grounding, and expected response quality."},
        {"field": "dspy_keyword_score", "description": f"Checks whether required expected keywords appear in the answer. Current threshold: {metric_thresholds['dspy_keyword_score']}", "importance": "Confirms the answer addressed the expected topic or intent."},
        {"field": "dspy_fallback_score", "description": f"Checks whether fallback behavior is correct. Current threshold: {metric_thresholds['dspy_fallback_score']}", "importance": "Helps detect non-answers or incorrect fallback handling."},
        {"field": "dspy_format_score", "description": f"Checks formatting constraints and answer cleanliness. Current threshold: {metric_thresholds['dspy_format_score']}", "importance": "Important for readability and consistency in customer-facing responses."},
        {"field": "dspy_pdf_grounding_score", "description": f"Checks grounding against the expected PDF/document source. Current threshold: {metric_thresholds['dspy_pdf_grounding_score']}", "importance": "Important for auditability and source correctness."},
        {"field": "ragas_answer_relevancy", "description": f"Measures semantic relevance of the answer to the question. Current threshold: {metric_thresholds['ragas_answer_relevancy']}", "importance": "Validates that the response is on-topic."},
        {"field": "ragas_answer_accuracy", "description": f"NVIDIA-style LLM judge metric that compares the response against the reference answer. Current threshold: {metric_thresholds['ragas_answer_accuracy']}", "importance": "Useful for checking how closely the answer matches the expected ground truth."},
        {"field": "ragas_faithfulness", "description": f"Measures whether the answer stays faithful to the retrieved context. Current threshold: {metric_thresholds['ragas_faithfulness']}", "importance": "Useful for hallucination detection when context is available."},
        {"field": "ragas_context_precision", "description": f"Measures whether retrieved context chunks are relevant and ranked well. Current threshold: {metric_thresholds['ragas_context_precision']}", "importance": "Useful for checking retriever quality and ranking."},
        {"field": "ragas_context_utilization", "description": f"Measures whether the retrieved context is actually useful for generating the response. Current threshold: {metric_thresholds['ragas_context_utilization']}", "importance": "Useful when you want to judge whether the answer really used the retrieved evidence."},
        {"field": "ragas_context_recall", "description": f"Measures whether important supporting information was not missed in retrieval. Current threshold: {metric_thresholds['ragas_context_recall']}", "importance": "Useful for checking that the retriever did not miss important content."},
        {"field": "ragas_context_relevance", "description": f"NVIDIA-style metric that judges whether the retrieved contexts are relevant to the user input. Current threshold: {metric_thresholds['ragas_context_relevance']}", "importance": "Useful for a lightweight overall relevance check on retrieval quality."},
        {"field": "ragas_response_groundedness", "description": f"NVIDIA-style metric that checks how well the response is grounded in the retrieved context. Current threshold: {metric_thresholds['ragas_response_groundedness']}", "importance": "Useful for a token-efficient groundedness check similar to faithfulness."},
        {"field": "ragas_context_entity_recall", "description": f"Measures how many important entities from the reference are covered by the retrieved context. Current threshold: {metric_thresholds['ragas_context_entity_recall']}", "importance": "Useful for fact-heavy scenarios where retrieving the right named entities matters."},
        {"field": "ragas_noise_sensitivity_relevant", "description": f"Measures how often incorrect claims appear when using retrieved relevant context. Lower is better; pass if <= {metric_thresholds['ragas_noise_sensitivity_relevant']}", "importance": "Useful for checking whether the system is being misled into wrong answers even when relevant evidence is present."},
        {"field": "ragas_noise_sensitivity_irrelevant", "description": f"Measures how often incorrect claims appear due to irrelevant/noisy context. Lower is better; pass if <= {metric_thresholds['ragas_noise_sensitivity_irrelevant']}", "importance": "Useful for checking robustness against distractor or noisy retrieval."},
        {"field": "ragas_response_correctness", "description": f"Combined semantic similarity + factual correctness vs ground truth (weighted F1). Current threshold: {metric_thresholds['ragas_response_correctness']}", "importance": "Measures overall correctness of the response without requiring retrieved contexts."},
        {"field": "ragas_answer_completeness", "description": f"LLM judge rates how fully the response addresses all aspects of the question (1-5 scale). Current threshold: {metric_thresholds['ragas_answer_completeness']}", "importance": "Useful for detecting partially-answered questions without needing retrieved contexts."},
        {"field": "Green / Red / Yellow", "description": "Green = pass, Red = fail, Yellow = not available / skipped.", "importance": "Allows instant visual scanning of metric health."},
        {"field": "artifact_json_path", "description": "Path to the raw chatbot evidence JSON.", "importance": "Provides traceability for audit and root-cause analysis."},
        {"field": "failure_screenshot_path", "description": "Path to Playwright screenshot captured on failure.", "importance": "Visual UI evidence for failed test cases."},
        {"field": "failure_trace_path", "description": "Path to Playwright trace.zip captured on failure.", "importance": "Detailed click/network/DOM replay for root-cause analysis."},
    ]


def _style_workbook(excel_path: Path) -> None:
    workbook = load_workbook(excel_path)

    for sheet in workbook.worksheets:
        if sheet.max_row >= 1:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_index, column_cells in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 14), 80)

    if "Test Results" in workbook.sheetnames:
        results_sheet = workbook["Test Results"]
        headers = {str(results_sheet.cell(row=1, column=index).value or ""): index for index in range(1, results_sheet.max_column + 1)}

        for row_index in range(2, results_sheet.max_row + 1):
            if "test_id" in headers:
                results_sheet.cell(row=row_index, column=headers["test_id"]).fill = BLUE_FILL

            if "final_status" in headers:
                status_cell = results_sheet.cell(row=row_index, column=headers["final_status"])
                if status_cell.value == "PASS":
                    status_cell.fill = GREEN_FILL
                elif status_cell.value == "FAIL":
                    status_cell.fill = RED_FILL

            for wrap_header in ["prompt", "answer_text", "ground_truth", "artifact_json_path", "failure_screenshot_path", "failure_trace_path"]:
                if wrap_header in headers:
                    results_sheet.cell(row=row_index, column=headers[wrap_header]).alignment = Alignment(wrap_text=True, vertical="top")

            for header, column_index in headers.items():
                if header.endswith("_result"):
                    cell = results_sheet.cell(row=row_index, column=column_index)
                    if cell.value == "PASS":
                        cell.fill = GREEN_FILL
                    elif cell.value == "FAIL":
                        cell.fill = RED_FILL
                    else:
                        cell.fill = YELLOW_FILL
                elif header.endswith("_pass_if>=") or header.endswith("_pass_if<="):
                    results_sheet.cell(row=row_index, column=column_index).fill = HEADER_FILL
                elif f"{header}_pass_if>=" in headers or f"{header}_pass_if<=" in headers:
                    score_cell = results_sheet.cell(row=row_index, column=column_index)
                    threshold_header = f"{header}_pass_if>=" if f"{header}_pass_if>=" in headers else f"{header}_pass_if<="
                    threshold_cell = results_sheet.cell(row=row_index, column=headers[threshold_header])
                    try:
                        score_value = _read_float(score_cell.value, 0.0)
                        threshold_value = _read_float(threshold_cell.value, 0.0)
                        if threshold_header.endswith("<="):
                            score_cell.fill = GREEN_FILL if score_value <= threshold_value else RED_FILL
                        else:
                            score_cell.fill = GREEN_FILL if score_value >= threshold_value else RED_FILL
                    except Exception:
                        if score_cell.value in (None, ""):
                            score_cell.fill = YELLOW_FILL

    if "Test Summary" in workbook.sheetnames:
        summary_sheet = workbook["Test Summary"]
        headers = {str(summary_sheet.cell(row=1, column=index).value or ""): index for index in range(1, summary_sheet.max_column + 1)}
        for row_index in range(2, summary_sheet.max_row + 1):
            if "PASS" in headers:
                summary_sheet.cell(row=row_index, column=headers["PASS"]).fill = GREEN_FILL
            if "FAIL" in headers:
                summary_sheet.cell(row=row_index, column=headers["FAIL"]).fill = RED_FILL
            if "N/A" in headers:
                summary_sheet.cell(row=row_index, column=headers["N/A"]).fill = YELLOW_FILL

    if "PDF Coverage" in workbook.sheetnames:
        coverage_sheet = workbook["PDF Coverage"]
        for cell in coverage_sheet["C"][1:]:
            if cell.value == "YES":
                cell.fill = GREEN_FILL
            elif cell.value == "NO":
                cell.fill = YELLOW_FILL

    if "Failures Only" in workbook.sheetnames:
        failures_sheet = workbook["Failures Only"]
        for row in failures_sheet.iter_rows(min_row=2, max_row=failures_sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    if "Legend" in workbook.sheetnames:
        legend_sheet = workbook["Legend"]
        for row in legend_sheet.iter_rows(min_row=2, max_row=legend_sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(excel_path)


def generate_enterprise_excel_report(
    project_root: str | Path,
    run_root: Path,
    dspy_results: dict[str, Any] | None = None,
    ragas_results: dict[str, Any] | None = None,
) -> Path:
    root = Path(project_root)
    reports_dir = root / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    # Latest_Report.xlsx is a temporary intermediate file — _sync_canonical_outputs_to_reports
    # copies it into the correct subfolder (ui_e2e) and it can then be cleaned up.
    excel_path = reports_dir / "Latest_Report.xlsx"

    ui_artifacts = _load_ui_artifacts(run_root / "ui_runs")
    dspy_payload = dspy_results or _load_json(run_root / "dspy" / "dspy_results.json")
    ragas_payload = ragas_results or _load_json(run_root / "ragas" / "ragas_results.json")

    results_rows, failures_rows, test_summary_rows, metric_thresholds = _build_results_rows(root, ui_artifacts, dspy_payload, ragas_payload)
    pdf_coverage_rows = _build_pdf_coverage_rows(root, results_rows)
    summary_rows = _build_summary_rows(results_rows, pdf_coverage_rows, run_root.name)
    legend_rows = _build_legend_rows(metric_thresholds)

    results_columns = [
        "test_id",
        "prompt",
        "run_id",
        "run_id_visible_in_prompt",
        "run_id_nonce_present",
        "answer_text",
        "ground_truth",
        "final_status",
        "dspy_total_score",
        "dspy_total_score_pass_if>=",
        "dspy_total_score_result",
        "dspy_keyword_score",
        "dspy_keyword_score_pass_if>=",
        "dspy_keyword_score_result",
        "dspy_fallback_score",
        "dspy_fallback_score_pass_if>=",
        "dspy_fallback_score_result",
        "dspy_format_score",
        "dspy_format_score_pass_if>=",
        "dspy_format_score_result",
        "dspy_pdf_grounding_score",
        "dspy_pdf_grounding_score_pass_if>=",
        "dspy_pdf_grounding_score_result",
        "ragas_answer_relevancy",
        "ragas_answer_relevancy_pass_if>=",
        "ragas_answer_relevancy_result",
        "ragas_answer_accuracy",
        "ragas_answer_accuracy_pass_if>=",
        "ragas_answer_accuracy_result",
        "ragas_faithfulness",
        "ragas_faithfulness_pass_if>=",
        "ragas_faithfulness_result",
        "ragas_context_precision",
        "ragas_context_precision_pass_if>=",
        "ragas_context_precision_result",
        "ragas_context_utilization",
        "ragas_context_utilization_pass_if>=",
        "ragas_context_utilization_result",
        "ragas_context_recall",
        "ragas_context_recall_pass_if>=",
        "ragas_context_recall_result",
        "ragas_context_relevance",
        "ragas_context_relevance_pass_if>=",
        "ragas_context_relevance_result",
        "ragas_response_groundedness",
        "ragas_response_groundedness_pass_if>=",
        "ragas_response_groundedness_result",
        "ragas_context_entity_recall",
        "ragas_context_entity_recall_pass_if>=",
        "ragas_context_entity_recall_result",
        "ragas_noise_sensitivity_relevant",
        "ragas_noise_sensitivity_relevant_pass_if<=",
        "ragas_noise_sensitivity_relevant_result",
        "ragas_noise_sensitivity_irrelevant",
        "ragas_noise_sensitivity_irrelevant_pass_if<=",
        "ragas_noise_sensitivity_irrelevant_result",
        "ragas_response_correctness",
        "ragas_response_correctness_pass_if>=",
        "ragas_response_correctness_result",
        "ragas_answer_completeness",
        "ragas_answer_completeness_pass_if>=",
        "ragas_answer_completeness_result",
        "expected_pdfs",
        "matched_pdfs",
        "artifact_json_path",
        "failure_screenshot_path",
        "failure_trace_path",
    ]

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Extract DSPy evaluation metadata for authentication
        dspy_metadata = dspy_payload.get("metadata", {}) if isinstance(dspy_payload, dict) else {}
        metadata_rows = []
        if dspy_metadata:
            metadata_rows.append({"attribute": "Evaluation Timestamp", "value": dspy_metadata.get("evaluation_timestamp", "N/A")})
            metadata_rows.append({"attribute": "LLM Provider", "value": dspy_metadata.get("llm_provider", "N/A")})
            metadata_rows.append({"attribute": "DSPy Version", "value": dspy_metadata.get("dspy_version", "N/A")})
            metadata_rows.append({"attribute": "Min Score Threshold", "value": dspy_metadata.get("min_score_threshold", "N/A")})
            metadata_rows.append({"attribute": "RAGAS Profile", "value": dspy_metadata.get("ragas_profile", "N/A")})
            metadata_rows.append({"attribute": "Results Path", "value": dspy_metadata.get("artifacts_path", "N/A")})
            metadata_rows.append({"attribute": "Results File", "value": dspy_metadata.get("dspy_results_file", "N/A")})
        
        if metadata_rows:
            pd.DataFrame(metadata_rows, columns=["attribute", "value"]).to_excel(
                writer,
                sheet_name="Evaluation_Metadata",
                index=False,
            )
        
        pd.DataFrame(test_summary_rows, columns=["evaluator", "pass_rule", "pass_threshold", "PASS", "FAIL", "N/A"]).to_excel(
            writer,
            sheet_name="Test Summary",
            index=False,
        )
        pd.DataFrame(results_rows, columns=results_columns).to_excel(
            writer,
            sheet_name="Test Results",
            index=False,
        )
        pd.DataFrame(pdf_coverage_rows, columns=["pdf_name", "total_tests", "tested"]).to_excel(
            writer,
            sheet_name="PDF Coverage",
            index=False,
        )
        pd.DataFrame(
            failures_rows,
            columns=["test_id", "failure_reason", "answer_text_excerpt", "artifact_json_path", "failure_screenshot_path", "failure_trace_path"],
        ).to_excel(
            writer,
            sheet_name="Failures Only",
            index=False,
        )
        pd.DataFrame(summary_rows, columns=["metric", "value"]).to_excel(
            writer,
            sheet_name="Summary Dashboard",
            index=False,
        )
        pd.DataFrame(legend_rows, columns=["field", "description", "importance"]).to_excel(
            writer,
            sheet_name="Legend",
            index=False,
        )

    _style_workbook(excel_path)
    return excel_path


def create_enterprise_reporting_assets(
    project_root: str | Path,
    dspy_results: dict[str, Any] | None = None,
    ragas_results: dict[str, Any] | None = None,
    report_mode: str = "all",
) -> dict[str, Any]:
    root = Path(project_root)
    execution_timestamp = datetime.now(timezone.utc).strftime(RUN_FOLDER_FORMAT)
    run_root = root / "artifacts" / "reports" / execution_timestamp
    run_root.mkdir(parents=True, exist_ok=True)

    _copy_tree(root / "artifacts" / "ui_runs", run_root / "ui_runs")
    _copy_tree(root / "artifacts" / "dspy", run_root / "dspy")
    _copy_tree(root / "artifacts" / "ragas", run_root / "ragas")
    _copy_file_if_exists(root / "artifacts" / "reports" / "pytest_report.html", run_root / "system" / "pytest_report.html")
    _copy_file_if_exists(root / "artifacts" / "reports" / "compliance_audit.json", run_root / "system" / "compliance_audit.json")

    excel_path = generate_enterprise_excel_report(root, run_root, dspy_results=dspy_results, ragas_results=ragas_results)

    _sync_canonical_outputs_to_reports(root, report_mode=report_mode)

    manifest = {
        "execution_timestamp": execution_timestamp,
        "run_root": str(run_root),
        "latest_excel_report": str(excel_path),
        "report_mode": report_mode,
        "traceability_note": "Real chatbot responses are preserved in artifacts/reports/<timestamp>/ui_runs/<test_id>.json and mirrored in artifacts/ui_runs/<test_id>.json.",
    }
    (run_root / "run_manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")

    # Keep report artifacts compact while preserving the latest run and canonical outputs.
    _apply_artifact_retention_policy(root)

    return manifest
