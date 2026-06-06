"""
verify_excel_report.py
======================
Confirms the bridge Excel report was created and lists its sheets.

Run:
    python verify_excel_report.py
"""

from __future__ import annotations
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
EXCEL = ROOT / "reports" / "bridge" / "Bridge_Evaluation_Report.xlsx"


def main() -> int:
    print("=" * 70)
    print("Bridge Excel report verifier")
    print("=" * 70)

    if not EXCEL.exists():
        print(f"[FAIL] Not found: {EXCEL}")
        return 2

    size = EXCEL.stat().st_size
    print(f"[OK]   File exists: {EXCEL}")
    print(f"[OK]   Size: {size:,} bytes")

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[WARN] openpyxl not installed; skipping content check.")
        return 0

    wb = load_workbook(EXCEL, read_only=True)
    print(f"[OK]   Sheets ({len(wb.sheetnames)}):")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"         - {name:<22}  rows={ws.max_row:>4}  cols={ws.max_column:>3}")

    if "Test_Summary" in wb.sheetnames:
        ws = wb["Test_Summary"]
        print("\n[Test_Summary preview]")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            print(f"  {row}")
            if i >= 14:
                break

    print("\n[OK]   Bridge Excel report is healthy. ✅")
    return 0


if __name__ == "__main__":
    sys.exit(main())