from __future__ import annotations

import os
from pathlib import Path

import pytest


PROJECT_ROOT = Path(__file__).resolve().parents[1]
UI_ARTIFACT_DIR = PROJECT_ROOT / "artifacts" / "ui_runs"
DSPY_OUTPUT_DIR = PROJECT_ROOT / "artifacts" / "dspy"
RAGAS_OUTPUT_DIR = PROJECT_ROOT / "artifacts" / "ragas"
ARCHIVED_REPORTS_DIR = PROJECT_ROOT / "artifacts" / "reports"
LATEST_REPORT_XLSX = PROJECT_ROOT / "reports" / "Latest_Report.xlsx"


@pytest.mark.ragas
def test_run_ragas_eval() -> None:
    pytest.importorskip("dspy", reason="DSPy is not installed in the active environment.")
    pytest.importorskip("ragas", reason="RAGAS is not installed in the active environment.")

    from dspy_layer.ui_to_dspy import convert_ui_artifacts_to_dspy_examples, run_dspy_evaluation
    from ragas_layer.dspy_to_ragas import convert_dspy_predictions_to_ragas_dataset
    from ragas_layer.ragas_runner import run_ragas_evaluation

    if not UI_ARTIFACT_DIR.exists() or not any(UI_ARTIFACT_DIR.glob("*.json")):
        pytest.skip("Run the UI capture suite first so artifacts are available.")

    examples = convert_ui_artifacts_to_dspy_examples(UI_ARTIFACT_DIR)
    if not examples:
        pytest.skip("No UI artifacts were found in artifacts/ui_runs.")

    dspy_results = run_dspy_evaluation(examples, output_path=DSPY_OUTPUT_DIR)
    dataset = convert_dspy_predictions_to_ragas_dataset(dspy_results)

    results = run_ragas_evaluation(
        dataset,
        metrics_config={
            "answer_relevancy_threshold": float(os.getenv("RAGAS_ANSWER_RELEVANCY_THRESHOLD", "0.65")),
            "faithfulness_threshold": float(os.getenv("RAGAS_FAITHFULNESS_THRESHOLD", "0.70")),
        },
        output_dir=RAGAS_OUTPUT_DIR,
    )

    assert (RAGAS_OUTPUT_DIR / "ragas_results.json").exists(), "RAGAS results JSON was not created."
    assert LATEST_REPORT_XLSX.exists(), "Enterprise Excel report `reports/Latest_Report.xlsx` was not created."

    from openpyxl import load_workbook

    workbook = load_workbook(LATEST_REPORT_XLSX, read_only=True)
    expected_sheets = {"Test Summary", "Test Results", "PDF Coverage", "Failures Only", "Summary Dashboard", "Legend"}
    assert expected_sheets.issubset(set(workbook.sheetnames)), (
        f"Latest report is missing expected sheets. Found: {workbook.sheetnames}"
    )

    report_headers = {cell.value for cell in next(workbook["Test Results"].iter_rows(min_row=1, max_row=1))}
    expected_metric_headers = {
        "ragas_answer_accuracy",
        "ragas_context_relevance",
        "ragas_response_groundedness",
        "ragas_context_entity_recall",
        "ragas_noise_sensitivity_relevant",
        "ragas_noise_sensitivity_irrelevant",
    }
    assert expected_metric_headers.issubset(report_headers), (
        f"Latest report is missing expected metric columns. Found headers: {sorted(str(h) for h in report_headers)}"
    )

    archived_run_dirs = [path for path in ARCHIVED_REPORTS_DIR.iterdir() if path.is_dir()]
    assert archived_run_dirs, "No timestamped report folder was created under `artifacts/reports/`."

    latest_run_dir = max(archived_run_dirs, key=lambda path: path.stat().st_mtime)
    assert (latest_run_dir / "ui_runs").exists(), "Timestamped UI evidence folder is missing."
    assert (latest_run_dir / "dspy" / "dspy_results.json").exists(), "Timestamped DSPy results JSON is missing."
    assert (latest_run_dir / "ragas" / "ragas_results.json").exists(), "Timestamped RAGAS results JSON is missing."

    executed = set(results.get("executed_metrics", []))
    skipped = {item["metric"]: item["reason"] for item in results.get("skipped_metrics", [])}
    summary = results.get("summary", {})
    has_contexts = any(bool(item) for item in dataset["contexts"])
    grounding_audit = results.get("document_grounding_audit", {})

    answer_relevancy_threshold = float(os.getenv("RAGAS_ANSWER_RELEVANCY_THRESHOLD", "0.65"))
    faithfulness_threshold = float(os.getenv("RAGAS_FAITHFULNESS_THRESHOLD", "0.70"))

    assert "answer_accuracy" in executed or "answer_accuracy" in skipped, (
        "Answer accuracy should be either executed or explicitly skipped."
    )

    assert grounding_audit.get("wrong_document_count", 0) == 0, (
        f"Wrong-document grounding detected: {grounding_audit.get('wrong_document_cases', [])}"
    )

    if "answer_relevancy" in executed:
        assert summary.get("answer_relevancy", 0.0) >= answer_relevancy_threshold, (
            f"answer_relevancy {summary.get('answer_relevancy')} is below threshold {answer_relevancy_threshold}."
        )
    else:
        pytest.skip(skipped.get("answer_relevancy", "answer_relevancy was skipped."))

    if has_contexts:
        if "faithfulness" in executed:
            assert summary.get("faithfulness", 0.0) >= faithfulness_threshold, (
                f"faithfulness {summary.get('faithfulness')} is below threshold {faithfulness_threshold}."
            )
        else:
            pytest.skip(skipped.get("faithfulness", "faithfulness was skipped."))
    else:
        assert "faithfulness" in skipped, "Faithfulness should be skipped when contexts are unavailable."
        assert "context_precision" in skipped, "Context precision should be skipped when contexts are unavailable."
        assert "context_utilization" in skipped, "Context utilization should be skipped when contexts are unavailable."
        assert "context_recall" in skipped, "Context recall should be skipped when contexts are unavailable."
        assert "context_relevance" in skipped, "Context relevance should be skipped when contexts are unavailable."
        assert "response_groundedness" in skipped, "Response groundedness should be skipped when contexts are unavailable."
        assert "context_entity_recall" in skipped, "Context entity recall should be skipped when contexts are unavailable."
        assert "noise_sensitivity_relevant" in skipped, "Noise sensitivity (relevant) should be skipped when contexts are unavailable."
        assert "noise_sensitivity_irrelevant" in skipped, "Noise sensitivity (irrelevant) should be skipped when contexts are unavailable."
