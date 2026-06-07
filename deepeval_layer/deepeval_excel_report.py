"""
deepeval_excel_report.py
========================
Enterprise-grade Excel report generator for DeepEval native evaluations.

Produces a multi-sheet Excel workbook with:
  Sheet 1 — Executive Summary  : KPIs, evaluator info, overall pass rate
  Sheet 2 — Metric Scorecard   : Per-metric avg, pass/fail, threshold, trend
  Sheet 3 — Test Case Detail   : Row-level scores for every case × metric
  Sheet 4 — Span & Trace Data  : LLM span timings, token counts, latency
  Sheet 5 — Failure Analysis   : Failed cases with reason, metric, suggestion
  Sheet 6 — Metric Correlation : Pairwise correlation matrix across metrics

Usage:
    from deepeval_layer.deepeval_excel_report import generate_deepeval_report
    path = generate_deepeval_report(results, output_dir=Path("reports/deepeval"))
    # or from CLI:
    python -m deepeval_layer.deepeval_excel_report --input eval_history/test_run_deepeval_<ts>.json
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))


def _require(pkg: str, install: str | None = None):
    try:
        return __import__(pkg)
    except ImportError:
        hint = install or pkg
        raise ImportError(f"Missing dependency: pip install {hint}")


# ── Excel styling constants ───────────────────────────────────────────────────

_HEADER_BG   = "1F497D"   # dark navy
_SECTION_BG  = "4472C4"   # medium blue
_PASS_BG     = "C6EFCE"   # light green
_FAIL_BG     = "FFC7CE"   # light red
_WARN_BG     = "FFEB9C"   # light yellow
_ALT_ROW_BG  = "EEF2F8"   # very light blue
_DEEPEVAL_BG = "7C3AED"   # purple (DeepEval brand)
_BORDER_COL  = "BFBFBF"

_FW_COLORS = {
    "ragas":         "6366F1",
    "azure-foundry": "10B981",
    "dspy":          "F59E0B",
    "deepeval":      "7C3AED",
}


def _cell_styles(ws, openpyxl_module):
    """Return style helpers bound to openpyxl."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    def fill(hex_color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def bold_font(size: int = 11, color: str = "000000", white: bool = False) -> Font:
        col = "FFFFFF" if white else color
        return Font(bold=True, size=size, color=col)

    def thin_border() -> Border:
        thin = Side(style="thin", color=_BORDER_COL)
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def center() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    return {"fill": fill, "bold": bold_font, "border": thin_border, "center": center, "left": left,
            "letter": get_column_letter}


def _auto_col_width(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col in ws.columns:
        col_letter = col[0].column_letter
        lengths = [len(str(cell.value or "")) for cell in col]
        w = max(min_w, min(max_w, max(lengths) + 2))
        ws.column_dimensions[col_letter].width = w


def _score_fill(score: float | None, threshold: float = 0.5):
    """Return a hex fill color based on score vs threshold."""
    from openpyxl.styles import PatternFill
    if score is None:
        return None
    if score >= threshold:
        return PatternFill(fill_type="solid", fgColor=_PASS_BG)
    if score >= threshold * 0.8:
        return PatternFill(fill_type="solid", fgColor=_WARN_BG)
    return PatternFill(fill_type="solid", fgColor=_FAIL_BG)


# ── Sheet 1: Executive Summary ────────────────────────────────────────────────

def _write_summary_sheet(ws, results: list[dict], run_meta: dict, sty):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = sty["fill"]; bold = sty["bold"]; border = sty["border"]
    center = sty["center"]; left = sty["left"]

    all_metrics = sorted({m["name"] for r in results for m in r.get("metrics", [])})
    total  = len(results)
    passed = sum(1 for r in results if r.get("success", False))
    failed = total - passed
    pass_rate = round(passed / total * 100, 1) if total else 0
    avg_cost  = round(sum(
        sum(m.get("evaluation_cost", 0.0) for m in r.get("metrics", []))
        for r in results
    ) / total, 6) if total else 0.0

    # Title block
    ws.merge_cells("A1:G1")
    ws["A1"] = "DeepEval Enterprise Evaluation Report — PlayReady DRM"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=_DEEPEVAL_BG)
    ws["A1"].alignment = center()

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Framework: DeepEval Native  |  Project: {run_meta.get('project','playready')}"
    ws["A2"].font = Font(size=10, color="6B7280")
    ws["A2"].alignment = center()

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18
    ws.append([])

    # KPI row
    kpis = [
        ("Total Test Cases", total,     None),
        ("Passed",           passed,    _PASS_BG),
        ("Failed",           failed,    _FAIL_BG if failed > 0 else None),
        ("Pass Rate",        f"{pass_rate}%", _PASS_BG if pass_rate >= 80 else _WARN_BG if pass_rate >= 60 else _FAIL_BG),
        ("Metrics Evaluated",len(all_metrics), None),
        ("Avg Cost / Case",  f"${avg_cost:.4f}", None),
    ]
    ws.append([k[0] for k in kpis])
    header_row = ws.max_row
    ws.append([k[1] for k in kpis])
    value_row = ws.max_row
    for col_idx, (label, val, bg) in enumerate(kpis, start=1):
        hc = ws.cell(row=header_row, column=col_idx)
        hc.font = Font(bold=True, size=10, color="FFFFFF")
        hc.fill = PatternFill(fill_type="solid", fgColor="374151")
        hc.alignment = center()
        vc = ws.cell(row=value_row, column=col_idx)
        vc.font = Font(bold=True, size=14)
        vc.alignment = center()
        if bg:
            vc.fill = PatternFill(fill_type="solid", fgColor=bg)

    ws.append([])

    # Run metadata table
    ws.append(["Run Metadata", ""])
    meta_row = ws.max_row
    ws.cell(meta_row, 1).font = bold(12, white=True)
    ws.cell(meta_row, 1).fill = fill("374151")
    meta_items = [
        ("Model",        run_meta.get("model", "gpt-4o")),
        ("Environment",  run_meta.get("environment", "production")),
        ("Version",      run_meta.get("version", "1.0.0")),
        ("Bot Type",     run_meta.get("bot_type", "—")),
        ("Evaluator",    "DeepEval Native"),
        ("Framework Key","deepeval"),
        ("Metrics",      ", ".join(all_metrics)),
    ]
    for k, v in meta_items:
        ws.append([k, str(v)])
        r = ws.max_row
        ws.cell(r, 1).font = bold(10)
        ws.cell(r, 1).fill = fill(_ALT_ROW_BG)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50


# ── Sheet 2: Metric Scorecard ─────────────────────────────────────────────────

def _write_metric_scorecard(ws, results: list[dict], sty):
    from openpyxl.styles import Font, PatternFill
    from collections import defaultdict
    fill = sty["fill"]; bold = sty["bold"]; border = sty["border"]; center = sty["center"]

    # Aggregate per metric
    acc: dict[str, dict] = defaultdict(lambda: {"scores": [], "passes": 0, "fails": 0, "threshold": 0.5, "reasons": []})
    for r in results:
        for m in r.get("metrics", []):
            n = m["name"]
            if m.get("score") is not None:
                acc[n]["scores"].append(m["score"])
            acc[n]["threshold"] = m.get("threshold", 0.5)
            if m.get("success"):
                acc[n]["passes"] += 1
            else:
                acc[n]["fails"] += 1
            if m.get("reason"):
                acc[n]["reasons"].append(m["reason"])

    headers = ["Metric", "Threshold", "Avg Score", "Best", "Worst", "Std Dev",
               "Pass", "Fail", "Pass Rate %", "Status", "Sample Reason"]
    ws.append(headers)
    hrow = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hrow, col)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill(fill_type="solid", fgColor=_DEEPEVAL_BG)
        c.alignment = center()

    import statistics

    for mname in sorted(acc.keys()):
        d = acc[mname]
        scores = d["scores"]
        avg  = round(sum(scores) / len(scores), 4) if scores else None
        best = round(max(scores), 4) if scores else None
        wrst = round(min(scores), 4) if scores else None
        std  = round(statistics.stdev(scores), 4) if len(scores) > 1 else 0.0
        thresh = d["threshold"]
        total_m = d["passes"] + d["fails"]
        pr = round(d["passes"] / total_m * 100, 1) if total_m else 0
        status = "PASS" if pr >= 80 else "WARN" if pr >= 60 else "FAIL"
        reason_sample = (d["reasons"][0] if d["reasons"] else "")[:120]

        ws.append([mname, thresh, avg, best, wrst, std, d["passes"], d["fails"], f"{pr}%", status, reason_sample])
        row_idx = ws.max_row
        sf = _score_fill(avg, thresh)
        if sf:
            ws.cell(row_idx, 3).fill = sf
        status_fill = _PASS_BG if status == "PASS" else _WARN_BG if status == "WARN" else _FAIL_BG
        ws.cell(row_idx, 10).fill = PatternFill(fill_type="solid", fgColor=status_fill)
        ws.cell(row_idx, 10).font = Font(bold=True)
        ws.cell(row_idx, 10).alignment = center()

    _auto_col_width(ws)


# ── Sheet 3: Test Case Detail ─────────────────────────────────────────────────

def _write_case_detail(ws, results: list[dict], sty):
    from openpyxl.styles import Font, PatternFill
    fill = sty["fill"]; bold = sty["bold"]; center = sty["center"]; left = sty["left"]

    all_metrics = sorted({m["name"] for r in results for m in r.get("metrics", [])})
    headers = ["#", "Name", "Result", "Question (truncated)", "Answer (truncated)"] + all_metrics
    ws.append(headers)
    hrow = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hrow, col)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor="1F497D")
        c.alignment = center()

    for i, r in enumerate(results, 1):
        success = r.get("success", False)
        result_label = "PASS" if success else "FAIL"
        score_map = {m["name"]: m for m in r.get("metrics", [])}
        row_data = [
            i,
            str(r.get("name", ""))[:60],
            result_label,
            str(r.get("question", ""))[:100],
            str(r.get("answer", ""))[:100],
        ] + [
            round(score_map[m]["score"], 3) if m in score_map and score_map[m].get("score") is not None else "—"
            for m in all_metrics
        ]
        ws.append(row_data)
        row_idx = ws.max_row
        # Alternate row color
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PatternFill(fill_type="solid", fgColor=_ALT_ROW_BG)
        # Result cell color
        res_cell = ws.cell(row_idx, 3)
        res_cell.fill = PatternFill(fill_type="solid", fgColor=_PASS_BG if success else _FAIL_BG)
        res_cell.font = Font(bold=True)
        res_cell.alignment = center()
        # Per-metric score cell colors
        for col_off, m in enumerate(all_metrics):
            col_idx = 6 + col_off
            cell = ws.cell(row_idx, col_idx)
            if m in score_map and score_map[m].get("score") is not None:
                thresh = score_map[m].get("threshold", 0.5)
                sf = _score_fill(score_map[m]["score"], thresh)
                if sf:
                    cell.fill = sf
            cell.alignment = center()

    # Freeze top row
    ws.freeze_panes = "A2"
    _auto_col_width(ws, max_w=40)


# ── Sheet 4: Span & Trace Data ────────────────────────────────────────────────

def _write_span_trace(ws, results: list[dict], sty):
    from openpyxl.styles import Font, PatternFill
    bold = sty["bold"]; center = sty["center"]

    headers = ["Case Name", "Trace Status", "Trace UUID (prefix)", "LLM Spans",
               "Retriever Spans", "LLM Model", "Input Tokens (est)", "Output Tokens (est)",
               "Cost/Case ($)", "Start Time", "End Time", "Latency (s)"]
    ws.append(headers)
    hrow = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hrow, col)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=_DEEPEVAL_BG)
        c.alignment = center()

    for i, r in enumerate(results, 1):
        trace = r.get("trace") or (r.get("traces") or [None])[0]
        if trace is None:
            # no trace info available
            ws.append([r.get("name", "")[:60]] + ["—"] * (len(headers) - 1))
            continue

        llm_spans = trace.get("llmSpans") or []
        ret_spans = trace.get("retrieverSpans") or []

        llm_s = llm_spans[0] if llm_spans else {}
        model = llm_s.get("model", "—")
        in_tok = llm_s.get("inputTokenCount", "—")
        out_tok = llm_s.get("outputTokenCount", "—")
        cpi = llm_s.get("costPerInputToken") or 0.0
        cpo = llm_s.get("costPerOutputToken") or 0.0
        cost = round((cpi * (in_tok if isinstance(in_tok, (int, float)) else 0))
                     + (cpo * (out_tok if isinstance(out_tok, (int, float)) else 0)), 6)
        start = llm_s.get("startTime", "—")
        end   = llm_s.get("endTime", "—")

        # Estimate latency from startTime / endTime strings
        latency = "—"
        try:
            from datetime import datetime as _dt
            fmt = "%Y-%m-%dT%H:%M:%S.%f"
            s = _dt.fromisoformat(str(start).replace("Z",""))
            e = _dt.fromisoformat(str(end).replace("Z",""))
            latency = round((e - s).total_seconds(), 3)
        except Exception:
            pass

        uuid_prefix = str(trace.get("uuid", ""))[:8]
        status = trace.get("status", "OK")

        ws.append([
            str(r.get("name", ""))[:60],
            status,
            uuid_prefix,
            len(llm_spans),
            len(ret_spans),
            model,
            in_tok,
            out_tok,
            cost if cost > 0 else "—",
            str(start)[:22],
            str(end)[:22],
            latency,
        ])
        row_idx = ws.max_row
        if i % 2 == 0:
            from openpyxl.styles import PatternFill as PF
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).fill = PF(fill_type="solid", fgColor=_ALT_ROW_BG)
        # Colour status
        stat_cell = ws.cell(row_idx, 2)
        stat_cell.fill = PatternFill(fill_type="solid",
                                     fgColor=_PASS_BG if status == "OK" else _FAIL_BG)
        stat_cell.alignment = center()

    ws.freeze_panes = "A2"
    _auto_col_width(ws)


# ── Sheet 5: Failure Analysis ─────────────────────────────────────────────────

def _write_failure_analysis(ws, results: list[dict], sty):
    from openpyxl.styles import Font, PatternFill
    bold = sty["bold"]; center = sty["center"]

    headers = ["#", "Case Name", "Failed Metrics", "Lowest Score", "Reason",
               "Suggested Action", "Question (truncated)"]
    ws.append(headers)
    hrow = ws.max_row
    for col, h in enumerate(headers, 1):
        c = ws.cell(hrow, col)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor="C0392B")
        c.alignment = center()

    SUGGESTIONS = {
        "PlayReadyRelevance":     "Review answer scope — ensure it addresses the specific PlayReady question.",
        "PlayReadyFaithfulness":  "Check retrieval context coverage; answer may reference undocumented claims.",
        "PlayReadyClarity":       "Restructure answer with clear headings; use correct PlayReady terminology.",
        "AnswerRelevancyMetric":  "Answer may be off-topic; improve retrieval to fetch more relevant chunks.",
        "FaithfulnessMetric":     "Reduce hallucination; constrain output to retrieved context.",
        "HallucinationMetric":    "High hallucination detected; add explicit grounding instructions to the prompt.",
    }

    fail_idx = 0
    for r in results:
        if r.get("success", True):
            continue
        fail_idx += 1
        failed_metrics = [m for m in r.get("metrics", []) if not m.get("success", True)]
        metric_names = ", ".join(m["name"] for m in failed_metrics)
        lowest = min((m["score"] for m in failed_metrics if m.get("score") is not None), default=None)
        reason = (failed_metrics[0].get("reason") or "")[:200] if failed_metrics else "—"
        primary_metric = failed_metrics[0]["name"] if failed_metrics else ""
        suggestion = SUGGESTIONS.get(primary_metric, "Review evaluation context and prompt instructions.")

        ws.append([
            fail_idx,
            str(r.get("name", ""))[:60],
            metric_names,
            round(lowest, 3) if lowest is not None else "—",
            reason,
            suggestion,
            str(r.get("question", ""))[:100],
        ])
        row_idx = ws.max_row
        # Score cell color
        if lowest is not None:
            ws.cell(row_idx, 4).fill = PatternFill(fill_type="solid", fgColor=_FAIL_BG)

    if fail_idx == 0:
        ws.append(["", "No failures detected — all cases passed.", "", "", "", "", ""])

    ws.freeze_panes = "A2"
    _auto_col_width(ws, max_w=50)


# ── Sheet 6: Metric Correlation ───────────────────────────────────────────────

def _write_correlation(ws, results: list[dict], sty):
    from openpyxl.styles import Font, PatternFill
    bold = sty["bold"]; center = sty["center"]

    all_metrics = sorted({m["name"] for r in results for m in r.get("metrics", [])})
    if not all_metrics:
        ws.append(["No metrics available"])
        return

    # Build score vectors
    score_vectors: dict[str, list[float]] = {m: [] for m in all_metrics}
    for r in results:
        sm = {m["name"]: m.get("score") for m in r.get("metrics", [])}
        for m in all_metrics:
            score_vectors[m].append(sm.get(m) or 0.0)

    # Header row
    ws.append(["Metric \\ Metric"] + all_metrics)
    hrow = ws.max_row
    for col, h in enumerate(["Metric \\ Metric"] + all_metrics, 1):
        c = ws.cell(hrow, col)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=_DEEPEVAL_BG)
        c.alignment = center()

    def _pearson(a, b):
        n = len(a)
        if n < 2:
            return None
        ma = sum(a) / n; mb = sum(b) / n
        num = sum((x - ma) * (y - mb) for x, y in zip(a, b))
        da = sum((x - ma) ** 2 for x in a) ** 0.5
        db = sum((y - mb) ** 2 for y in b) ** 0.5
        if da == 0 or db == 0:
            return None
        return round(num / (da * db), 3)

    for row_m in all_metrics:
        row_data = [row_m]
        for col_m in all_metrics:
            corr = _pearson(score_vectors[row_m], score_vectors[col_m])
            row_data.append(corr if corr is not None else "—")
        ws.append(row_data)
        row_idx = ws.max_row
        ws.cell(row_idx, 1).font = Font(bold=True)
        for col_off, col_m in enumerate(all_metrics):
            cell = ws.cell(row_idx, 2 + col_off)
            val = cell.value
            if isinstance(val, float):
                if val >= 0.8:
                    cell.fill = PatternFill(fill_type="solid", fgColor=_PASS_BG)
                elif val >= 0.5:
                    cell.fill = PatternFill(fill_type="solid", fgColor=_WARN_BG)
                elif val < 0:
                    cell.fill = PatternFill(fill_type="solid", fgColor=_FAIL_BG)
            cell.alignment = center()

    _auto_col_width(ws)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_deepeval_report(
    results: list[dict],
    output_dir: Path | None = None,
    run_meta: dict | None = None,
) -> Path:
    """
    Generate the enterprise Excel report.

    Args:
        results:    List of result row dicts (same format as deepeval_to_dashboard).
        output_dir: Directory to write to (default: reports/deepeval/).
        run_meta:   Optional dict with model, environment, version, bot_type, project.

    Returns:
        Path to the written .xlsx file.
    """
    openpyxl = _require("openpyxl")
    from openpyxl import Workbook

    if output_dir is None:
        output_dir = PROJECT_ROOT / "reports" / "deepeval"
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    if run_meta is None:
        run_meta = {"model": "gpt-4o", "project": "playready", "environment": "production", "version": "1.0.0"}

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheets = [
        ("Summary",     lambda ws: _write_summary_sheet(ws, results, run_meta, None)),
        ("Scorecard",   lambda ws: _write_metric_scorecard(ws, results, None)),
        ("Case Detail", lambda ws: _write_case_detail(ws, results, None)),
        ("Spans",       lambda ws: _write_span_trace(ws, results, None)),
        ("Failures",    lambda ws: _write_failure_analysis(ws, results, None)),
        ("Correlation", lambda ws: _write_correlation(ws, results, None)),
    ]

    for name, writer in sheets:
        ws = wb.create_sheet(title=name)
        try:
            # We need to pass sty — create it per sheet
            import openpyxl as opx
            sty = _cell_styles(ws, opx)
            if name == "Summary":
                _write_summary_sheet(ws, results, run_meta, sty)
            elif name == "Scorecard":
                _write_metric_scorecard(ws, results, sty)
            elif name == "Case Detail":
                _write_case_detail(ws, results, sty)
            elif name == "Spans":
                _write_span_trace(ws, results, sty)
            elif name == "Failures":
                _write_failure_analysis(ws, results, sty)
            elif name == "Correlation":
                _write_correlation(ws, results, sty)
        except Exception as e:
            ws.append([f"Error generating sheet: {e}"])

    from datetime import datetime as _dt, timezone as _tz
    ts = _dt.now(_tz.utc).strftime("%Y%m%d_%H%M%S")
    fname = output_dir / f"deepeval_enterprise_report_{ts}.xlsx"
    wb.save(str(fname))
    print(f"  [deepeval-report] Wrote enterprise Excel -> {fname}")
    return fname


def generate_deepeval_report_from_json(json_path: Path, output_dir: Path | None = None) -> Path:
    """Load a test_run_deepeval_*.json and generate the Excel report from it."""
    raw = json.loads(Path(json_path).read_text(encoding="utf-8"))
    test_cases = raw.get("testCases") or []
    hyper = raw.get("hyperparameters") or {}

    results = []
    for tc in test_cases:
        metrics_raw = tc.get("metricsData") or []
        metrics = []
        for m in metrics_raw:
            metrics.append({
                "name":            m.get("name", ""),
                "score":           m.get("score"),
                "threshold":       m.get("threshold", 0.5),
                "success":         m.get("success", False),
                "reason":          m.get("reason", ""),
                "evaluation_cost": m.get("evaluationCost", 0.0),
                "model":           m.get("evaluationModel", ""),
            })
        trace = tc.get("trace")
        results.append({
            "name":     tc.get("name", ""),
            "question": tc.get("input", ""),
            "answer":   tc.get("actualOutput", ""),
            "contexts": tc.get("retrievalContext") or [],
            "success":  tc.get("success", False),
            "metrics":  metrics,
            "trace":    trace,
        })

    run_meta = {
        "model":       hyper.get("model", "gpt-4o"),
        "project":     hyper.get("project", "playready"),
        "environment": hyper.get("environment", "production"),
        "version":     hyper.get("version", "1.0.0"),
        "bot_type":    hyper.get("bot_type", "—"),
    }
    return generate_deepeval_report(results, output_dir=output_dir, run_meta=run_meta)


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Generate DeepEval enterprise Excel report")
    ap.add_argument("--input", type=Path, required=True,
                    help="Path to test_run_deepeval_*.json or deepeval_results.json")
    ap.add_argument("--output-dir", type=Path, default=None,
                    help="Output directory (default: reports/deepeval/)")
    args = ap.parse_args()

    dest = generate_deepeval_report_from_json(args.input, args.output_dir)
    print(f"Report: {dest}")
