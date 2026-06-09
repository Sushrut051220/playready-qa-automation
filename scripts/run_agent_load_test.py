from __future__ import annotations

import argparse
import csv
import json
import math
import os
import statistics
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from azure.identity import DefaultAzureCredential
from azure.ai.projects import AIProjectClient

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# PROJECT SETUP
# =========================================================
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / ".env", override=True)

os.environ.pop("AZURE_OPENAI_API_KEY", None)


# =========================================================
# INDUSTRY / QA SLA BENCHMARKS
# =========================================================
SLA_TARGETS = {
    "avg_latency_seconds": 10.0,
    "p50_latency_seconds": 8.0,
    "p95_latency_seconds": 12.0,
    "failure_rate_percent": 5.0,
    "throughput_req_per_sec": 1.0,
    "avg_tokens_per_request": 8000.0,
}


# =========================================================
# CLIENT
# =========================================================
def build_client():
    endpoint = os.environ["FOUNDRY_PROJECT_ENDPOINT"]

    project_client = AIProjectClient(
        endpoint=endpoint,
        credential=DefaultAzureCredential(),
    )

    openai_client = project_client.get_openai_client()
    agent_name = os.environ.get("FOUNDRY_AGENT_NAME", "PublicAgent")
    agent_version = os.environ.get("FOUNDRY_AGENT_VERSION", "11")

    return openai_client, agent_name, agent_version


# =========================================================
# RESPONSE HELPERS
# =========================================================
def extract_answer_and_citations(response):
    answer_text = ""
    citations = []
    citation_quotes = []

    for item in (getattr(response, "output", None) or []):
        if getattr(item, "type", "") != "message":
            continue

        for content_item in (getattr(item, "content", None) or []):
            text_value = getattr(content_item, "text", "") or ""
            if text_value:
                answer_text += text_value

            for ann in (getattr(content_item, "annotations", None) or []):
                ann_type = getattr(ann, "type", "")
                if ann_type in {"url_citation", "file_citation"}:
                    citation_entry = {
                        "type": ann_type,
                        "url": getattr(ann, "url", ""),
                        "title": getattr(ann, "title", ""),
                        "text": getattr(ann, "text", "") or "",
                        "file_id": getattr(ann, "file_id", ""),
                        "filename": getattr(ann, "filename", ""),
                        "start_index": getattr(ann, "start_index", 0),
                        "end_index": getattr(ann, "end_index", 0),
                    }
                    citations.append(citation_entry)

                    quote = (
                        citation_entry["title"]
                        or citation_entry["filename"]
                        or citation_entry["text"]
                    )
                    if quote:
                        citation_quotes.append(quote)

    return answer_text, citations, citation_quotes


def percentile(values, p):
    if not values:
        return 0.0
    if len(values) == 1:
        return float(values[0])

    values_sorted = sorted(values)
    rank = (len(values_sorted) - 1) * (p / 100.0)
    lower = math.floor(rank)
    upper = math.ceil(rank)

    if lower == upper:
        return float(values_sorted[int(rank)])

    weight = rank - lower
    return float(values_sorted[lower] * (1 - weight) + values_sorted[upper] * weight)


# =========================================================
# DATA LOADING
# =========================================================
def load_queries_from_json(input_path, limit=0, offset=0):
    raw_cases = json.loads(input_path.read_text(encoding="utf-8-sig"))
    if not isinstance(raw_cases, list) or not raw_cases:
        raise ValueError("Input JSON is empty or not a JSON array")

    if offset:
        raw_cases = raw_cases[offset:]
    if limit:
        raw_cases = raw_cases[:limit]
    return raw_cases


def build_query_pool(raw_cases, repeat):
    pool = []
    for cycle in range(repeat):
        for i, case in enumerate(raw_cases, start=1):
            question = case.get("prompt") or case.get("question") or ""
            case_id = case.get("id", f"case_{i}")
            pool.append({
                "request_id": f"{case_id}_r{cycle+1}",
                "case_id": case_id,
                "question": question,
            })
    return pool


# =========================================================
# AGENT CALL
# =========================================================
def _iso_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


def query_agent(client, agent_name, agent_version, question, request_id):
    start_time = time.time()
    start_iso = _iso_now()
    try:
        response = client.responses.create(
            input=[{"role": "user", "content": question}],
            extra_body={
                "agent_reference": {
                    "name": agent_name,
                    "version": agent_version,
                    "type": "agent_reference",
                }
            },
        )
    except Exception as e:
        end_iso = _iso_now()
        return {
            "request_id": request_id,
            "question": question,
            "answer": f"[AGENT CALL FAILED: {type(e).__name__}: {str(e)[:500]}]",
            "citations": [],
            "citation_quotes": [],
            "run_status": "failed",
            "latency_seconds": round(time.time() - start_time, 4),
            "request_start_iso": start_iso,
            "request_end_iso": end_iso,
            "token_usage": {"total_tokens": 0, "prompt_tokens": 0, "completion_tokens": 0},
            "error_type": type(e).__name__,
            "error_message": str(e)[:1000],
            "answer_length": 0,
            "citation_count": 0,
        }

    latency = round(time.time() - start_time, 4)
    end_iso = _iso_now()
    usage = getattr(response, "usage", None)
    token_usage = {
        "total_tokens": getattr(usage, "total_tokens", 0),
        "prompt_tokens": getattr(usage, "input_tokens", 0),
        "completion_tokens": getattr(usage, "output_tokens", 0),
    }

    answer_text, citations, citation_quotes = extract_answer_and_citations(response)

    return {
        "request_id": request_id,
        "question": question,
        "answer": answer_text,
        "citations": citations,
        "citation_quotes": citation_quotes,
        "run_status": "completed",
        "latency_seconds": latency,
        "request_start_iso": start_iso,
        "request_end_iso": end_iso,
        "token_usage": token_usage,
        "error_type": "",
        "error_message": "",
        "answer_length": len(answer_text or ""),
        "citation_count": len(citations),
    }


# =========================================================
# REPORTING HELPERS
# =========================================================
def flatten_result(result):
    token_usage = result.get("token_usage", {}) or {}
    citations = result.get("citations", []) or []
    citation_quotes = result.get("citation_quotes", []) or []

    return {
        "request_id": result.get("request_id", ""),
        "question": result.get("question", ""),
        "run_status": result.get("run_status", ""),
        "latency_seconds": result.get("latency_seconds", 0.0),
        "answer_length": result.get("answer_length", 0),
        "citation_count": result.get("citation_count", 0),
        "citation_quotes": " | ".join(citation_quotes),
        "total_tokens": token_usage.get("total_tokens", 0),
        "prompt_tokens": token_usage.get("prompt_tokens", 0),
        "completion_tokens": token_usage.get("completion_tokens", 0),
        "error_type": result.get("error_type", ""),
        "error_message": result.get("error_message", ""),
        "answer": result.get("answer", ""),
        "citations_json": json.dumps(citations, ensure_ascii=False),
    }


def evaluate_sla(summary):
    benchmark_targets = {
        "avg_latency_target_seconds": SLA_TARGETS["avg_latency_seconds"],
        "p50_target_seconds": SLA_TARGETS["p50_latency_seconds"],
        "p95_target_seconds": SLA_TARGETS["p95_latency_seconds"],
        "failure_rate_target_percent": SLA_TARGETS["failure_rate_percent"],
        "throughput_target_req_per_sec": SLA_TARGETS["throughput_req_per_sec"],
        "avg_tokens_target_per_request": SLA_TARGETS["avg_tokens_per_request"],
    }

    sla_status = {
        "avg_latency_status": "PASS" if summary["latency_avg_seconds"] <= SLA_TARGETS["avg_latency_seconds"] else "FAIL",
        "p50_status": "PASS" if summary["latency_p50_seconds"] <= SLA_TARGETS["p50_latency_seconds"] else "FAIL",
        "p95_status": "PASS" if summary["latency_p95_seconds"] <= SLA_TARGETS["p95_latency_seconds"] else "FAIL",
        "failure_rate_status": "PASS" if summary["failure_rate_percent"] <= SLA_TARGETS["failure_rate_percent"] else "FAIL",
        "throughput_status": "PASS" if summary["throughput_req_per_sec"] >= SLA_TARGETS["throughput_req_per_sec"] else "FAIL",
        "token_efficiency_status": "PASS" if summary["avg_tokens_per_request"] <= SLA_TARGETS["avg_tokens_per_request"] else "FAIL",
    }
    sla_status["overall_status"] = "PASS" if all(v == "PASS" for v in sla_status.values()) else "FAIL"
    return sla_status, benchmark_targets


def generate_rca(summary):
    rca = []

    if summary["latency_avg_seconds"] > SLA_TARGETS["avg_latency_seconds"]:
        rca.append("Average latency exceeds the dev-defined SLA of 10s, indicating overall slow response across users.")
    if summary["latency_p95_seconds"] > SLA_TARGETS["p95_latency_seconds"]:
        rca.append("P95 latency exceeds conversational RAG SLA threshold, indicating slow worst-case user experience.")
    if summary["avg_tokens_per_request"] > SLA_TARGETS["avg_tokens_per_request"]:
        rca.append("Average token usage per request is too high, indicating excessive retrieved context or prompt construction overhead.")
    if summary["throughput_req_per_sec"] < SLA_TARGETS["throughput_req_per_sec"]:
        rca.append("Observed throughput is below target, indicating limited scalability under concurrent usage.")
    if summary["failure_rate_percent"] > SLA_TARGETS["failure_rate_percent"]:
        rca.append("Failure rate exceeds acceptable threshold, indicating reliability issues under load.")
    if not rca:
        rca.append("No major performance bottleneck exceeded defined SLA thresholds in this run.")

    return rca


def generate_recommendations(summary, sla_status):
    recs = []

    if sla_status.get("avg_latency_status") == "FAIL":
        recs.append("Reduce average latency by optimizing prompt construction, context size, and model deployment.")
    if sla_status["p95_status"] == "FAIL":
        recs.append("Optimize RAG retrieval and orchestration to lower worst-case latency.")
    if sla_status["token_efficiency_status"] == "FAIL":
        recs.append("Reduce retrieved context size and remove redundant chunks to lower prompt token usage.")
    if sla_status["throughput_status"] == "FAIL":
        recs.append("Increase concurrency capacity and validate scaling with 5, 10, and 20 users.")
    if sla_status["failure_rate_status"] == "FAIL":
        recs.append("Investigate agent call failures and add retry / fallback mechanisms.")
    if not recs:
        recs.append("System meets SLA targets; continue staged load validation to confirm scaling behavior.")

    return recs


def generate_industry_verdict(summary, sla_status):
    if sla_status["overall_status"] == "PASS":
        return "PASS - Agent performance is within defined SLA thresholds for this run."
    return "FAIL - Agent performance does not meet one or more SLA thresholds and is not production-ready from a performance perspective."


def safe_to_excel(writer, df, sheet_name):
    if df is None or df.empty:
        df = pd.DataFrame([{"info": "No data available for this section."}])
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def apply_color_coding(excel_path):
    wb = load_workbook(excel_path)
    if not wb.sheetnames:
        return

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.sheet_state = "visible"

        if ws.max_row < 1 or ws.max_column < 1:
            ws["A1"] = "No data"
            continue

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        for col in range(1, ws.max_column + 1):
            max_len = 10
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[get_column_letter(col)].width = max_len + 2

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str):
                    val = cell.value.strip().upper()
                    if val == "PASS":
                        cell.fill = green
                    elif val == "FAIL":
                        cell.fill = red

    if not any(ws.sheet_state == "visible" for ws in wb.worksheets):
        wb.worksheets[0].sheet_state = "visible"

    wb.save(excel_path)

def build_test_summary_report(summary, sla_status, benchmark_targets, verdict, rca, recommendations):
    rows = []

    # 1. Test Info
    rows.append({"Section": "TEST INFORMATION", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Agent", "Value": f"{summary['agent_name']}:{summary['agent_version']}"})
    rows.append({"Section": "", "Item": "Bot Type", "Value": summary.get('bot_type', 'public')})
    rows.append({"Section": "", "Item": "Environment", "Value": summary['project_endpoint']})
    rows.append({"Section": "", "Item": "Input File", "Value": summary['input_file']})
    rows.append({"Section": "", "Item": "Test Date", "Value": time.strftime("%Y-%m-%d %H:%M:%S")})

    # 2. Configuration
    rows.append({"Section": "TEST CONFIGURATION", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Concurrent Users", "Value": summary['users']})
    rows.append({"Section": "", "Item": "Repeat Cycles", "Value": summary['repeat']})
    rows.append({"Section": "", "Item": "Source Test Cases", "Value": summary['source_case_count']})
    rows.append({"Section": "", "Item": "Total Requests Executed", "Value": summary['request_count']})

    # 3. Execution Results
    rows.append({"Section": "EXECUTION RESULTS", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Completed Requests", "Value": summary['completed_count']})
    rows.append({"Section": "", "Item": "Failed Requests", "Value": summary['failed_count']})
    rows.append({"Section": "", "Item": "Failure Rate (%)", "Value": summary['failure_rate_percent']})
    rows.append({"Section": "", "Item": "Total Duration (s)", "Value": summary['total_duration_seconds']})
    rows.append({"Section": "", "Item": "Throughput (req/sec)", "Value": summary['throughput_req_per_sec']})

    # 4. Performance
    rows.append({"Section": "PERFORMANCE METRICS", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Avg Latency (s)", "Value": summary['latency_avg_seconds']})
    rows.append({"Section": "", "Item": "P50 Latency (s)", "Value": summary['latency_p50_seconds']})
    rows.append({"Section": "", "Item": "P90 Latency (s)", "Value": summary['latency_p90_seconds']})
    rows.append({"Section": "", "Item": "P95 Latency (s)", "Value": summary['latency_p95_seconds']})
    rows.append({"Section": "", "Item": "P99 Latency (s)", "Value": summary['latency_p99_seconds']})
    rows.append({"Section": "", "Item": "Min Latency (s)", "Value": summary['latency_min_seconds']})
    rows.append({"Section": "", "Item": "Max Latency (s)", "Value": summary['latency_max_seconds']})

    # 5. Quality
    rows.append({"Section": "QUALITY INDICATORS", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Total Tokens Used", "Value": summary['total_tokens']})
    rows.append({"Section": "", "Item": "Avg Tokens / Request", "Value": summary['avg_tokens_per_request']})
    rows.append({"Section": "", "Item": "Avg Answer Length (chars)", "Value": summary['avg_answer_length_chars']})
    rows.append({"Section": "", "Item": "Avg Citations / Request", "Value": summary['avg_citations_per_request']})

    # 6. SLA Evaluation
    rows.append({"Section": "SLA EVALUATION", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Avg Latency Status", "Value": sla_status['avg_latency_status']})
    rows.append({"Section": "", "Item": "P50 Status", "Value": sla_status['p50_status']})
    rows.append({"Section": "", "Item": "P95 Status", "Value": sla_status['p95_status']})
    rows.append({"Section": "", "Item": "Failure Rate Status", "Value": sla_status['failure_rate_status']})
    rows.append({"Section": "", "Item": "Throughput Status", "Value": sla_status['throughput_status']})
    rows.append({"Section": "", "Item": "Token Efficiency Status", "Value": sla_status['token_efficiency_status']})

    # 7. Verdict
    rows.append({"Section": "FINAL VERDICT", "Item": "", "Value": ""})
    rows.append({"Section": "", "Item": "Overall Status", "Value": sla_status['overall_status']})
    rows.append({"Section": "", "Item": "Verdict", "Value": verdict})

    # 8. RCA
    rows.append({"Section": "ROOT CAUSE ANALYSIS", "Item": "", "Value": ""})
    for item in rca:
        rows.append({"Section": "", "Item": "RCA", "Value": item})

    # 9. Recommendations
    rows.append({"Section": "RECOMMENDATIONS", "Item": "", "Value": ""})
    for item in recommendations:
        rows.append({"Section": "", "Item": "Recommendation", "Value": item})

    return pd.DataFrame(rows)
# =========================================================
# MAIN LOAD TEST
# =========================================================
def run_load_test(input_path, output_dir, users, repeat, limit, offset, ramp_pause_seconds, bot_type="public"):
    print(f"Loading test cases from: {input_path}")

    raw_cases = load_queries_from_json(input_path, limit, offset)
    print(f"Loaded {len(raw_cases)} source test cases")

    pool = build_query_pool(raw_cases, repeat)
    print(f"Total requests to execute: {len(pool)}")
    print(f"Concurrent users: {users}")

    client, agent_name, agent_version = build_client()
    print(f"Foundry agent : {agent_name}:{agent_version}")
    print(f"Project       : {os.environ.get('FOUNDRY_PROJECT_ENDPOINT')}\n")

    all_results = []
    test_start = time.time()

    for batch_start in range(0, len(pool), users):
        batch = pool[batch_start: batch_start + users]
        batch_no = batch_start // users + 1
        print(f"--- Batch {batch_no} | requests {batch_start + 1} to {batch_start + len(batch)} ---")

        with ThreadPoolExecutor(max_workers=users) as executor:
            future_map = {
                executor.submit(
                    query_agent, client, agent_name, agent_version,
                    item["question"], item["request_id"]
                ): item for item in batch
            }

            for future in as_completed(future_map):
                item = future_map[future]
                try:
                    result = future.result()
                except Exception as e:
                    result = {
                        "request_id": item["request_id"],
                        "question": item["question"],
                        "answer": f"[UNEXPECTED FAILURE: {type(e).__name__}: {str(e)[:500]}]",
                        "citations": [],
                        "citation_quotes": [],
                        "run_status": "failed",
                        "latency_seconds": 0.0,
                        "token_usage": {"total_tokens": 0, "prompt_tokens": 0, "completion_tokens": 0},
                        "error_type": type(e).__name__,
                        "error_message": str(e)[:1000],
                        "answer_length": 0,
                        "citation_count": 0,
                    }

                all_results.append(result)
                print(
                    f"  {result['request_id']} | "
                    f"status={result['run_status']} | "
                    f"lat={result['latency_seconds']}s | "
                    f"tokens={result['token_usage']['total_tokens']} | "
                    f"chars={result['answer_length']} | "
                    f"citations={result['citation_count']}"
                )

        if ramp_pause_seconds > 0 and batch_start + users < len(pool):
            time.sleep(ramp_pause_seconds)

    total_seconds = round(time.time() - test_start, 4)

    latencies = [r["latency_seconds"] for r in all_results if r["latency_seconds"] > 0]
    failed = [r for r in all_results if r["run_status"] != "completed"]
    completed = [r for r in all_results if r["run_status"] == "completed"]
    total_tokens = sum(r["token_usage"]["total_tokens"] for r in all_results)
    prompt_tokens = sum(r["token_usage"]["prompt_tokens"] for r in all_results)
    completion_tokens = sum(r["token_usage"]["completion_tokens"] for r in all_results)

    summary = {
        "agent_name": agent_name,
        "agent_version": agent_version,
        "bot_type": bot_type,
        "project_endpoint": os.environ.get("FOUNDRY_PROJECT_ENDPOINT", ""),
        "input_file": str(input_path),
        "users": users,
        "repeat": repeat,
        "source_case_count": len(raw_cases),
        "request_count": len(all_results),
        "completed_count": len(completed),
        "failed_count": len(failed),
        "failure_rate_percent": round((len(failed) / len(all_results) * 100), 2) if all_results else 0.0,
        "total_duration_seconds": total_seconds,
        "throughput_req_per_sec": round(len(all_results) / total_seconds, 4) if total_seconds > 0 else 0.0,
        "latency_avg_seconds": round(statistics.mean(latencies), 4) if latencies else 0.0,
        "latency_median_seconds": round(statistics.median(latencies), 4) if latencies else 0.0,
        "latency_p50_seconds": round(percentile(latencies, 50), 4) if latencies else 0.0,
        "latency_p90_seconds": round(percentile(latencies, 90), 4) if latencies else 0.0,
        "latency_p95_seconds": round(percentile(latencies, 95), 4) if latencies else 0.0,
        "latency_p99_seconds": round(percentile(latencies, 99), 4) if latencies else 0.0,
        "latency_min_seconds": round(min(latencies), 4) if latencies else 0.0,
        "latency_max_seconds": round(max(latencies), 4) if latencies else 0.0,
        "total_tokens": total_tokens,
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "avg_tokens_per_request": round(total_tokens / len(all_results), 2) if all_results else 0.0,
        "avg_answer_length_chars": round(statistics.mean([r["answer_length"] for r in all_results]), 2) if all_results else 0.0,
        "avg_citations_per_request": round(statistics.mean([r["citation_count"] for r in all_results]), 2) if all_results else 0.0,
        "error_breakdown": {},
    }

    error_breakdown = {}
    for r in failed:
        key = r.get("error_type") or "UnknownError"
        error_breakdown[key] = error_breakdown.get(key, 0) + 1
    summary["error_breakdown"] = error_breakdown

    sla_status, benchmark_targets = evaluate_sla(summary)
    rca = generate_rca(summary)
    recommendations = generate_recommendations(summary, sla_status)
    verdict = generate_industry_verdict(summary, sla_status)

    output_dir.mkdir(parents=True, exist_ok=True)

    ts = time.strftime("%Y%m%d_%H%M%S")
    summary_path = output_dir / f"agent_load_summary_{ts}.json"
    details_json_path = output_dir / f"agent_load_details_{ts}.json"
    details_csv_path = output_dir / f"agent_load_details_{ts}.csv"
    excel_path = output_dir / f"agent_load_bridge_{ts}.xlsx"

    full_summary_package = {
        "summary": summary,
        "sla_status": sla_status,
        "benchmark_targets": benchmark_targets,
        "verdict": verdict,
        "root_cause_analysis": rca,
        "recommendations": recommendations,
    }

    summary_path.write_text(json.dumps(full_summary_package, indent=2, ensure_ascii=False), encoding="utf-8")
    details_json_path.write_text(json.dumps(all_results, indent=2, ensure_ascii=False), encoding="utf-8")

    flat_results = [flatten_result(r) for r in all_results]

    csv_fieldnames = list(flat_results[0].keys()) if flat_results else [
        "request_id", "question", "run_status", "latency_seconds",
        "answer_length", "citation_count", "citation_quotes",
        "total_tokens", "prompt_tokens", "completion_tokens",
        "error_type", "error_message", "answer", "citations_json"
    ]

    with details_csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fieldnames)
        writer.writeheader()
        for row in flat_results:
            writer.writerow(row)

    # =========================================================
    # Excel Report (FINAL — safe)
    # =========================================================
    summary_df = pd.DataFrame([summary])
    results_df = pd.DataFrame(flat_results) if flat_results else pd.DataFrame([{"info": "No results"}])
    latency_df = pd.DataFrame({"latency_seconds": latencies}) if latencies else pd.DataFrame([{"info": "No latency data"}])
    error_rows = [{"error_type": k, "count": v} for k, v in error_breakdown.items()]
    errors_df = pd.DataFrame(error_rows) if error_rows else pd.DataFrame([{"info": "No errors"}])
    benchmark_df = pd.DataFrame([benchmark_targets])
    sla_status_df = pd.DataFrame([sla_status])
    rca_df = pd.DataFrame({"Root Cause Analysis": rca})
    recommendations_df = pd.DataFrame({"Recommendations": recommendations})
    verdict_df = pd.DataFrame([{"Final Verdict": verdict}])

    test_report_df = build_test_summary_report(summary, sla_status, benchmark_targets, verdict, rca, recommendations)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        safe_to_excel(writer, test_report_df, "Test Report")
        safe_to_excel(writer, summary_df, "Summary")
        safe_to_excel(writer, sla_status_df, "Evaluation")
        safe_to_excel(writer, benchmark_df, "SLA Benchmarks")
        safe_to_excel(writer, verdict_df, "Final Verdict")
        safe_to_excel(writer, rca_df, "RCA")
        safe_to_excel(writer, recommendations_df, "Recommendations")
        safe_to_excel(writer, results_df, "Detailed Results")
        safe_to_excel(writer, latency_df, "Latency Details")
        safe_to_excel(writer, errors_df, "Errors")

    # ✅ Apply color AFTER Excel saved
    apply_color_coding(excel_path)

    print("\n==================================================")
    print("INDUSTRY-GRADE LOAD TEST SUMMARY")
    print("==================================================")
    print(f"Agent                 : {summary['agent_name']}:{summary['agent_version']}")
    print(f"Requests              : {summary['request_count']}")
    print(f"Completed             : {summary['completed_count']}")
    print(f"Failed                : {summary['failed_count']}")
    print(f"Failure rate          : {summary['failure_rate_percent']}% ({sla_status['failure_rate_status']})")
    print(f"Total duration        : {summary['total_duration_seconds']}s")
    print(f"Throughput            : {summary['throughput_req_per_sec']} req/s ({sla_status['throughput_status']})")
    print(f"Avg latency           : {summary['latency_avg_seconds']}s ({sla_status['avg_latency_status']})")
    print(f"P50 latency           : {summary['latency_p50_seconds']}s ({sla_status['p50_status']})")
    print(f"P95 latency           : {summary['latency_p95_seconds']}s ({sla_status['p95_status']})")
    print(f"Min latency           : {summary['latency_min_seconds']}s")
    print(f"Max latency           : {summary['latency_max_seconds']}s")
    print(f"Total tokens          : {summary['total_tokens']}")
    print(f"Avg tokens / request  : {summary['avg_tokens_per_request']} ({sla_status['token_efficiency_status']})")
    print(f"Overall Status        : {sla_status['overall_status']}")
    print("--------------------------------------------------")
    print("Root Cause Analysis:")
    for item in rca:
        print(f"  - {item}")
    print("--------------------------------------------------")
    print("Recommendations:")
    for item in recommendations:
        print(f"  - {item}")
    print("--------------------------------------------------")
    print(f"Summary JSON          : {summary_path}")
    print(f"Details JSON          : {details_json_path}")
    print(f"Details CSV           : {details_csv_path}")
    print(f"Bridge Excel Report   : {excel_path}")
    print("==================================================")

    return full_summary_package, all_results, details_json_path, summary_path


# =========================================================
# ENTRY
# =========================================================
def main():
    parser = argparse.ArgumentParser(
        description="Run PlayReady agent load test and optionally push results to DeepEval dashboard."
    )
    parser.add_argument("--input",         default=str(PROJECT_ROOT / "data" / "test_cases.json"))
    parser.add_argument("--output-dir",    default=str(PROJECT_ROOT / "reports" / "load_testing"))
    parser.add_argument("--users",         type=int,   default=5)
    parser.add_argument("--repeat",        type=int,   default=1)
    parser.add_argument("--limit",         type=int,   default=10)
    parser.add_argument("--offset",        type=int,   default=0)
    parser.add_argument("--ramp-pause",    type=float, default=0.0)
    parser.add_argument("--bot-type",      default="public", choices=["public", "customer", "private"])
    parser.add_argument(
        "--dashboard-dir",
        default=None,
        help="Path to the DeepEval dashboard eval_history folder. "
             "When provided the results are automatically pushed to the dashboard "
             "after the load test completes (single-command workflow).",
    )

    args = parser.parse_args()

    try:
        result = run_load_test(
            input_path=Path(args.input),
            output_dir=Path(args.output_dir),
            users=args.users,
            repeat=args.repeat,
            limit=args.limit,
            offset=args.offset,
            ramp_pause_seconds=args.ramp_pause,
            bot_type=args.bot_type,
        )
        _, _, details_path, summary_path = result

        if args.dashboard_dir:
            print("\n[Dashboard] Pushing results to dashboard …")
            from scripts.load_test_to_dashboard import convert as _lt_convert
            out = _lt_convert(
                details_path=details_path,
                summary_path=summary_path,
                bot_type=args.bot_type,
                output_dir=args.dashboard_dir,
            )
            print(f"[Dashboard] Done → {out}")

    except Exception as e:
        print(f"FATAL: {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()