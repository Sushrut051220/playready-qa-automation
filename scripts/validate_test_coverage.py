from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"

REQUIRED_SCENARIOS = [
    {"id_suffix": "pos_factual", "category": "positive", "needs_chunk": True},
    {"id_suffix": "pos_detail", "category": "positive", "needs_chunk": True},
    {"id_suffix": "edge_version", "category": "edge_case", "needs_chunk": True},
    {"id_suffix": "edge_negation", "category": "edge_case", "needs_chunk": True},
    {"id_suffix": "edge_multipart", "category": "edge_case", "needs_chunk": True},
    {"id_suffix": "neg_out_of_scope", "category": "negative", "needs_chunk": False},
    {"id_suffix": "neg_wrong_doc", "category": "negative", "needs_chunk": False},
    {"id_suffix": "halluc_false_premise", "category": "hallucination", "needs_chunk": False},
    {"id_suffix": "halluc_nonexistent", "category": "hallucination", "needs_chunk": True},
    {"id_suffix": "robust_paraphrase", "category": "robustness", "needs_chunk": True},
    {"id_suffix": "comp_rule", "category": "compliance", "needs_chunk": True},
    {"id_suffix": "ground_cite", "category": "grounding", "needs_chunk": True},
    {"id_suffix": "cross_doc", "category": "cross_document", "needs_chunk": True},
    {"id_suffix": "sec_pii_exposure", "category": "data_security", "needs_chunk": True},
    {"id_suffix": "sec_data_retention", "category": "data_security", "needs_chunk": True},
    {"id_suffix": "sec_encryption_keys", "category": "data_security", "needs_chunk": True},
    {"id_suffix": "sec_access_control", "category": "data_security", "needs_chunk": True},
    {"id_suffix": "sec_audit_logging", "category": "data_security", "needs_chunk": True},
    {"id_suffix": "conv_casual", "category": "conversational", "needs_chunk": True},
    {"id_suffix": "conv_frustrated", "category": "conversational", "needs_chunk": True},
]

REQUIRED_IDS = [s["id_suffix"] for s in REQUIRED_SCENARIOS]


def _load():
    p = DATA_DIR / "test_cases.json"
    return json.loads(p.read_text(encoding="utf-8-sig"))


def _scen_id(cid):
    parts = cid.split("_", 2)
    return parts[2] if len(parts) >= 3 else ""


def validate_coverage():
    cases = _load()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    pdf_map = {}
    globs = []
    for c in cases:
        cid = c.get("id", "")
        if cid.startswith("neg_") or cid.startswith("conv_"):
            globs.append(c)
        else:
            pdf = c.get("source_pdf") or c.get("expected_pdfs", ["unknown"])[0]
            pdf_map.setdefault(pdf, []).append(c)

    print(f"Loaded {len(cases)} test cases")
    print(f"  Per-PDF: {sum(len(v) for v in pdf_map.values())}")
    print(f"  Global: {len(globs)}")
    print(f"  Required per PDF: {len(REQUIRED_IDS)}")
    print(f"  PDFs: {len(pdf_map)}")
    print()

    cov_rows = []
    gap_rows = []
    pdf_rows = []
    tp = 0
    tf = 0

    for pn in sorted(pdf_map.keys()):
        pcs = pdf_map[pn]
        found = set()
        for c in pcs:
            sid = _scen_id(c.get("id", ""))
            if sid:
                found.add(sid)

        miss = [s for s in REQUIRED_IDS if s not in found]
        matched = found & set(REQUIRED_IDS)
        pct = round((len(matched) / len(REQUIRED_IDS)) * 100, 1)
        ok = len(miss) == 0

        if ok:
            tp += 1
        else:
            tf += 1

        pdf_rows.append({
            "source_pdf": pn, "total_cases": len(pcs),
            "scenarios_found": len(matched), "scenarios_required": len(REQUIRED_IDS),
            "coverage_pct": pct, "missing_count": len(miss),
            "missing_scenarios": ", ".join(miss) if miss else "",
            "status": "COMPLETE" if ok else "INCOMPLETE",
            "run_timestamp": ts,
        })

        for sc in REQUIRED_SCENARIOS:
            sid = sc["id_suffix"]
            f = sid in found
            hc = False
            hg = False
            hr = False
            if f:
                mc = next((x for x in pcs if _scen_id(x.get("id", "")) == sid), None)
                if mc:
                    hc = bool(mc.get("expected_chunk_ids"))
                    hg = bool(mc.get("ground_truth"))
                    hr = bool(mc.get("reference_contexts"))
            nc = sc["needs_chunk"]
            co = (not nc) or hc
            q = "FULL" if (f and co and hg) else "PARTIAL" if f else "MISSING"
            cov_rows.append({
                "source_pdf": pn, "scenario_id": sid, "category": sc["category"],
                "found": "YES" if f else "NO", "needs_chunk": "YES" if nc else "NO",
                "has_chunk": "YES" if hc else "NO", "has_gt": "YES" if hg else "NO",
                "has_ctx": "YES" if hr else "NO", "quality": q, "run_timestamp": ts,
            })
            if not f:
                gap_rows.append({"source_pdf": pn, "missing": sid, "category": sc["category"], "run_timestamp": ts})

        icon = "+" if ok else "X"
        print(f"  [{icon}] {pn}: {pct}% ({len(miss)} missing)")

    cat_s = {}
    for c in cases:
        ct = c.get("query_type", "unknown")
        cat_s.setdefault(ct, {"n": 0, "g": 0, "c": 0})
        cat_s[ct]["n"] += 1
        if c.get("ground_truth"):
            cat_s[ct]["g"] += 1
        if c.get("expected_chunk_ids"):
            cat_s[ct]["c"] += 1

    cat_rows = []
    for cn, st in sorted(cat_s.items()):
        cat_rows.append({
            "category": cn, "total": st["n"],
            "with_gt": st["g"],
            "gt_pct": round((st["g"] / st["n"]) * 100, 1) if st["n"] else 0,
            "with_chunks": st["c"],
            "chunk_pct": round((st["c"] / st["n"]) * 100, 1) if st["n"] else 0,
            "run_timestamp": ts,
        })

    ov = {
        "total_pdfs": len(pdf_map), "complete": tp, "incomplete": tf,
        "coverage_pct": round((tp / len(pdf_map)) * 100, 1) if pdf_map else 0,
        "total_cases": len(cases),
        "pdf_cases": sum(len(v) for v in pdf_map.values()),
        "neg_cases": len([c for c in globs if c.get("id", "").startswith("neg_")]),
        "conv_cases": len([c for c in globs if c.get("id", "").startswith("conv_")]),
        "gaps": len(gap_rows), "run_timestamp": ts,
    }

    print()
    print("=" * 60)
    print("OVERALL COVERAGE: " + str(ov["coverage_pct"]) + "%")
    print("  PDFs complete:   " + str(tp) + "/" + str(len(pdf_map)))
    print("  PDFs incomplete: " + str(tf) + "/" + str(len(pdf_map)))
    print("  Total cases:     " + str(len(cases)))
    print("  Per-PDF:         " + str(ov["pdf_cases"]))
    print("  Negative:        " + str(ov["neg_cases"]))
    print("  Conversational:  " + str(ov["conv_cases"]))
    print("  Gaps:            " + str(len(gap_rows)))
    print("=" * 60)

    from openpyxl.styles import PatternFill, Font, Alignment
    GF = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RF = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YF = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    HF = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HN = Font(color="FFFFFF", bold=True)

    def sh(ws):
        row1 = list(ws.iter_rows(min_row=1, max_row=1))[0]
        for cell in row1:
            cell.fill = HF
            cell.font = HN
            cell.alignment = Alignment(horizontal="center")
            cell.font = HN
            cell.alignment = Alignment(horizontal="center")

    def aw(ws):
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 60)

    def cc(ws, cn, pv, fv):
        ci = None
        for i, c in enumerate(ws[1], 1):
            if c.value == cn:
                ci = i
                break
        if ci:
            for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                for cell in row:
                    v = str(cell.value).upper() if cell.value else ""
                    if v == pv.upper() or v == "YES" or v == "FULL":
                        cell.fill = GF
                    elif v == fv.upper() or v == "NO" or v == "MISSING":
                        cell.fill = RF
                    elif v == "PARTIAL":
                        cell.fill = YF

    rd = PROJECT_ROOT / "reports" / "bridge"
    rd.mkdir(parents=True, exist_ok=True)
    rp = rd / "Test_Coverage_Report.xlsx"

    with pd.ExcelWriter(rp, engine="openpyxl") as w:
        pd.DataFrame([ov]).to_excel(w, sheet_name="Overall_Summary", index=False)
        sh(w.sheets["Overall_Summary"])
        aw(w.sheets["Overall_Summary"])

        pd.DataFrame(pdf_rows).to_excel(w, sheet_name="PDF_Coverage", index=False)
        ws = w.sheets["PDF_Coverage"]
        sh(ws)
        aw(ws)
        cc(ws, "status", "COMPLETE", "INCOMPLETE")

        pd.DataFrame(cov_rows).to_excel(w, sheet_name="Coverage_Matrix", index=False)
        ws = w.sheets["Coverage_Matrix"]
        sh(ws)
        aw(ws)
        cc(ws, "found", "YES", "NO")
        cc(ws, "quality", "FULL", "MISSING")

        if gap_rows:
            pd.DataFrame(gap_rows).to_excel(w, sheet_name="Gaps", index=False)
        else:
            pd.DataFrame([{"status": "No gaps - 100% coverage!"}]).to_excel(w, sheet_name="Gaps", index=False)
        sh(w.sheets["Gaps"])
        aw(w.sheets["Gaps"])

        pd.DataFrame(cat_rows).to_excel(w, sheet_name="Category_Summary", index=False)
        sh(w.sheets["Category_Summary"])
        aw(w.sheets["Category_Summary"])

        rr = []
        for i, s in enumerate(REQUIRED_SCENARIOS, 1):
            nc = "YES" if s["needs_chunk"] else "NO"
            rr.append({"num": i, "scenario": s["id_suffix"], "category": s["category"], "needs_chunk": nc})
        pd.DataFrame(rr).to_excel(w, sheet_name="Scenario_Reference", index=False)
        sh(w.sheets["Scenario_Reference"])
        aw(w.sheets["Scenario_Reference"])

    print()
    print("Report: " + str(rp))

    jp = DATA_DIR / "coverage_report.json"
    payload = {"overall": ov, "pdf_coverage": pdf_rows, "categories": cat_rows, "gaps": gap_rows}
    jp.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print("JSON: " + str(jp))
    return ov


if __name__ == "__main__":
    validate_coverage()